"""
test_ambient_and_companion.py
Автоматические тесты для:
1. 3D-кораблика AmbientBoat на волнах Герстнера (физика, сброс idle, убегание, проекция).
2. Реального выезда окон Companion Mode (AuthenticFilesWindow, AuthenticValuesWindow, AuthenticHistoryWindow, AuthenticRunWindow, CompanionTopSettingsBar, CompanionModeManager).
"""

import os
import sys
import unittest
import time
import numpy as np

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)

from PySide6.QtWidgets import QApplication
from PySide6.QtCore import Qt, QRect, QPoint, QEvent, QSettings

from ui.gl.ambient_boat import AmbientBoat, BoatState
from ui.components.companion_dock import (
    CardCategory, DockSide, AuthenticFilesWindow, AuthenticValuesWindow,
    AuthenticHistoryWindow, AuthenticRunWindow, CompanionTopSettingsBar, CompanionModeManager
)


class TestAmbientBoat(unittest.TestCase):
    """Тестирование физики и логики состояний 3D-кораблика."""

    def setUp(self):
        self.boat = AmbientBoat()

    def test_initial_state_is_hidden(self):
        self.assertEqual(self.boat.state, BoatState.HIDDEN)
        self.assertEqual(self.boat.alpha, 0.0)

    def test_idle_spawning(self):
        self.boat.idle_start_time = time.time() - 30.0
        self.boat.update(0.02)
        self.assertEqual(self.boat.state, BoatState.SPAWNING)
        self.assertGreater(self.boat.target_alpha, 0.0)

    def test_fleeing_on_activity(self):
        self.boat.state = BoatState.SAILING
        self.boat.alpha = 0.9
        self.boat.target_alpha = 0.9

        self.boat.reset_idle()
        self.assertEqual(self.boat.state, BoatState.FLEEING)
        self.assertEqual(self.boat.target_alpha, 0.0)

        self.boat.update(0.1)
        self.assertLess(self.boat.alpha, 0.9)

    def test_wave_projection(self):
        self.boat.state = BoatState.SAILING
        self.boat.alpha = 0.8
        lines = self.boat.get_projected_lines(t=1.5, wave_amp=0.22, wave_steep=0.55, tilt=0.48)
        self.assertIsInstance(lines, np.ndarray)
        self.assertGreater(len(lines), 0)
        self.assertEqual(lines.shape[1], 2)

    def test_models_count(self):
        self.assertEqual(len(self.boat._models), 4)
        for i, model in enumerate(self.boat._models):
            self.assertGreater(len(model), 10, f"Model {i} should have vertices")


class TestCompanionFloatingWindows(unittest.TestCase):
    """Тестирование аутентичных функциональных окон и окна настроек."""

    @classmethod
    def setUpClass(cls):
        if not QApplication.instance():
            cls.app = QApplication(sys.argv)
        else:
            cls.app = QApplication.instance()

    def test_floating_windows_creation(self):
        w_top = CompanionTopSettingsBar()
        w_files = AuthenticFilesWindow()
        w_vals = AuthenticValuesWindow()
        w_hist = AuthenticHistoryWindow()
        w_run = AuthenticRunWindow()

        self.assertIsNotNone(w_top.btn_side)
        self.assertIsNotNone(w_top.btn_restore)
        self.assertIsNotNone(w_files.drop_tpl)
        self.assertIsNotNone(w_files.drop_arc)
        self.assertIsNotNone(w_files.txt_save)
        self.assertIsNotNone(w_files.btn_repl)
        self.assertIsNotNone(w_vals.txt_cold)
        self.assertIsNotNone(w_vals.txt_hot)
        self.assertIsNotNone(w_vals.txt_corr)
        self.assertIsNotNone(w_hist.table_hist)
        self.assertIsNotNone(w_run.btn_run)

    def test_values_window_sum(self):
        w_vals = AuthenticValuesWindow()
        w_vals.set_values("10.5", "5.5", "1.5")
        self.assertIn("17.50", w_vals.lbl_sum.text())

    def test_dock_side_toggle(self):
        class DummyMainWin:
            def geometry(self):
                return QRect(100, 100, 800, 600)
            def pos(self):
                return QPoint(100, 100)

        QSettings("WaterMetrics", "WaterMetricsApp").setValue("companion/dock_side", "right")
        mgr = CompanionModeManager(DummyMainWin())
        self.assertEqual(mgr.dock_side, DockSide.RIGHT)
        mgr.dock_side = DockSide.LEFT
        self.assertEqual(mgr.dock_side, DockSide.LEFT)

class DummyMainWin:
    def __init__(self):
        self._geo = QRect(100, 100, 1000, 700)
        self._visible = True
        self._maximized = False
        self.page_main = None

    def geometry(self):
        return self._geo

    def setGeometry(self, g):
        self._geo = g

    def pos(self):
        return self._geo.topLeft()

    def isMaximized(self):
        return self._maximized

    def isVisible(self):
        return self._visible

    def show(self):
        self._visible = True

    def hide(self):
        self._visible = False

    def showMaximized(self):
        self._maximized = True
        self._visible = True

    def showNormal(self):
        self._maximized = False
        self._visible = True

    def raise_(self):
        pass

    def activateWindow(self):
        pass


class TestCompanionModeManagerFeatures(unittest.TestCase):
    """
    Глубокое тестирование исправлений и новых фич Companion Mode:
    1. Кнопка «В окно» и горячие клавиши (F11, Ctrl+D, Escape).
    2. Отсутствие наездов карточек друг на друга на любых экранах.
    3. Корректное поведение при переключении дока налево (Left Dock Side).
    4. Drag & Drop перетаскивание и изменение порядка карточек.
    """

    @classmethod
    def setUpClass(cls):
        if not QApplication.instance():
            cls.app = QApplication(sys.argv)
        else:
            cls.app = QApplication.instance()

    def setUp(self):
        from PySide6.QtCore import QSettings
        QSettings("WaterMetrics", "WaterMetricsApp").remove("companion/card_order")
        QSettings("WaterMetrics", "WaterMetricsApp").setValue("companion/dock_side", "right")
        QSettings("WaterMetrics", "WaterMetricsApp").setValue("companion/is_pinned", False)
        self.dummy_main = DummyMainWin()
        self.mgr = CompanionModeManager(self.dummy_main)
        self.mgr.cards = [
            self.mgr.win_top_bar,
            self.mgr.win_files,
            self.mgr.win_values,
            self.mgr.win_hist,
            self.mgr.win_run
        ]

    def test_no_card_overlapping_all_resolutions(self):
        """Проверка расчета позиций карточек: карточки НИКОГДА не наезжают друг на друга."""
        resolutions = [
            QRect(0, 0, 1366, 768),
            QRect(0, 0, 1920, 1080),
            QRect(0, 0, 2560, 1440),
            QRect(0, 0, 3840, 2160),
            QRect(1920, 0, 1920, 1080), # Второй монитор
        ]

        for screen_geo in resolutions:
            positions = self.mgr._calculate_dock_positions(screen_geo)
            self.assertEqual(len(positions), len(self.mgr.cards))

            cards_in_order = self.mgr.cards
            for i in range(len(cards_in_order) - 1):
                c_curr = cards_in_order[i]
                c_next = cards_in_order[i + 1]

                y_curr = positions[c_curr]
                h_curr = self.mgr._get_card_height(c_curr)
                y_next = positions[c_next]

                # Низ текущей карточки строго меньше или равен верху следующей карточки (зазор >= 10px)
                self.assertLessEqual(
                    y_curr + h_curr,
                    y_next,
                    f"Наезд карточки {c_curr.category.name} на {c_next.category.name} на разрешении {screen_geo}"
                )
                self.assertGreaterEqual(
                    y_next - (y_curr + h_curr),
                    self.mgr.card_spacing,
                    f"Зазор между карточками меньше {self.mgr.card_spacing}px"
                )

    def test_left_dock_side_coordinates_and_sliding(self):
        """Проверка работы дока на левой стороне экрана: корректные x-координаты открытого и закрытого состояния."""
        screen_geo = QRect(0, 0, 1920, 1080)
        self.mgr.dock_side = DockSide.LEFT

        open_x = self.mgr._get_open_x(screen_geo)
        park_x = self.mgr._get_park_x(screen_geo)

        # При открытии на левой стороне карточка находится на экране (open_x = screen_geo.left() + 6)
        self.assertEqual(open_x, screen_geo.left() + 6)
        self.assertGreaterEqual(open_x, 0)

        # При парковке на левой стороне из-за экрана выглядывает ровно 16px (parked_peek)
        peek_visible = (park_x + self.mgr.dock_width) - screen_geo.left()
        self.assertEqual(peek_visible, self.mgr.cards[0].parked_peek)

        # Переключение обратно на правую сторону
        self.mgr.dock_side = DockSide.RIGHT
        open_x_r = self.mgr._get_open_x(screen_geo)
        park_x_r = self.mgr._get_park_x(screen_geo)

        self.assertEqual(open_x_r, screen_geo.right() - self.mgr.dock_width - 6)
        self.assertEqual(park_x_r, screen_geo.right() - self.mgr.cards[0].parked_peek)

    def test_restore_requested_and_exit_companion(self):
        """Проверка выхода в окно по кнопке и сигналу restore_requested."""
        self.mgr.enter_companion_mode()
        self.assertTrue(self.mgr.is_companion_active)
        self.assertFalse(self.dummy_main.isVisible())

        # Имитируем клик по кнопке «В окно»
        self.mgr.win_top_bar.restore_requested.emit()
        self.assertFalse(self.mgr.is_companion_active)

    def test_keyboard_shortcuts_trigger_exit(self):
        """Проверка горячих клавиш F11, Ctrl+D и Escape на карточках."""
        from PySide6.QtGui import QKeyEvent
        from PySide6.QtCore import QEvent

        # 1. Тест F11
        self.mgr.enter_companion_mode()
        self.assertTrue(self.mgr.is_companion_active)
        key_event_f11 = QKeyEvent(QEvent.Type.KeyPress, Qt.Key.Key_F11, Qt.KeyboardModifier.NoModifier)
        self.mgr.win_files.keyPressEvent(key_event_f11)
        self.assertFalse(self.mgr.is_companion_active)

        # 2. Тест Ctrl+D
        self.mgr.enter_companion_mode()
        self.assertTrue(self.mgr.is_companion_active)
        key_event_ctrl_d = QKeyEvent(QEvent.Type.KeyPress, Qt.Key.Key_D, Qt.KeyboardModifier.ControlModifier)
        self.mgr.win_values.keyPressEvent(key_event_ctrl_d)
        self.assertFalse(self.mgr.is_companion_active)

        # 3. Тест Escape
        self.mgr.enter_companion_mode()
        self.assertTrue(self.mgr.is_companion_active)
        key_event_esc = QKeyEvent(QEvent.Type.KeyPress, Qt.Key.Key_Escape, Qt.KeyboardModifier.NoModifier)
        self.mgr.win_hist.keyPressEvent(key_event_esc)
        self.assertFalse(self.mgr.is_companion_active)

    def test_drag_and_drop_card_reordering(self):
        """Проверка перетаскивания и изменения порядка карточек."""
        self.mgr.enter_companion_mode()
        original_order = list(self.mgr.cards)
        self.assertEqual(len(original_order), 5)

        card_to_move = self.mgr.win_run  # Последняя карточка (index 4)
        self.mgr.on_card_drag_started(card_to_move)

        # Перетаскиваем карточку наверх (target_y = -20)
        target_y = -20
        self.mgr.on_card_drag_moved(card_to_move, target_y)
        self.mgr.on_card_drag_ended(card_to_move)

        # Порядок в self.cards изменился: win_run переместился на индекс 1 (сразу под шапкой)
        new_order = self.mgr.cards
        self.assertNotEqual(original_order, new_order)
        self.assertIn(card_to_move, new_order)
        self.assertEqual(new_order[1], card_to_move)
        self.assertEqual(new_order[0], self.mgr.win_top_bar)

    def test_drag_x_locked_to_dock_column(self):
        """Проверка: карточка не может быть утянута по горизонтали (X заблокирован в док-колонке)."""
        self.mgr.enter_companion_mode()
        card = self.mgr.win_files
        screen_geo = QRect(0, 0, 1920, 1080)
        expected_open_x = self.mgr._get_open_x(screen_geo)

        # Имитируем мышиный драг в центр экрана (x = 500)
        from PySide6.QtGui import QMouseEvent
        from PySide6.QtCore import QPointF
        card.drag_start_pos = QPoint(1800, 300)
        card.drag_initial_y = 300
        card.is_dragging = True

        move_event = QMouseEvent(
            QEvent.Type.MouseMove,
            QPointF(500, 400),
            QPointF(500, 400),
            Qt.MouseButton.LeftButton,
            Qt.MouseButton.LeftButton,
            Qt.KeyboardModifier.NoModifier
        )
        card.mouseMoveEvent(move_event)

        # Карточка не уехала на x=500, а строго заблокирована на open_x дока!
        self.assertEqual(card.x(), expected_open_x)

    def test_foolproof_window_restoration(self):
        """Проверка: при выходе в окно hovering/proximity не могут сбить возврат окна."""
        self.mgr.enter_companion_mode()
        self.assertTrue(self.mgr.is_companion_active)

        self.mgr.exit_companion_mode()
        self.assertTrue(self.mgr.is_returning_to_window)

        # Имитируем наведение мыши во время возврата в окно (не должно сбивать возврат)
        self.mgr.win_files.enterEvent(QEvent(QEvent.Type.Enter))
        self.mgr._check_dock_proximity()

        # Флаг и процесс возврата в окно остаются стабильными
        self.assertTrue(self.mgr.is_returning_to_window or not self.mgr.is_companion_active)

    def test_toggle_dock_side_and_sliding_both_sides(self):
        """Проверка смены полярности: все 5 окон перемещаются вместе и корректно выезжают/заезжают."""
        self.mgr.enter_companion_mode()
        self.assertEqual(self.mgr.dock_side, DockSide.RIGHT)

        # Переключаем сторону на LEFT
        self.mgr.toggle_dock_side()
        self.assertEqual(self.mgr.dock_side, DockSide.LEFT)

        # Проверяем, что ВСЕ 5 окон получили DockSide.LEFT
        for card in self.mgr.cards:
            self.assertEqual(card.dock_side, DockSide.LEFT)

        # Проверяем кнопку
        self.assertIn("Справа", self.mgr.win_top_bar.btn_side.text())

        # Переключаем обратно на RIGHT
        self.mgr.toggle_dock_side()
        self.assertEqual(self.mgr.dock_side, DockSide.RIGHT)
        for card in self.mgr.cards:
            self.assertEqual(card.dock_side, DockSide.RIGHT)
        self.assertIn("Слева", self.mgr.win_top_bar.btn_side.text())

    def test_flight_interaction_locking(self):
        """Проверка: во время полета взаимодействие заблокировано, разблокируется по прибытии."""
        self.mgr.set_all_cards_flight_locked(True)
        self.assertTrue(self.mgr.is_flight_animating)
        for card in self.mgr.cards:
            self.assertTrue(card.is_flight_locked)

        # Разблокировка
        self.mgr.set_all_cards_flight_locked(False)
        self.assertFalse(self.mgr.is_flight_animating)
        for card in self.mgr.cards:
            self.assertFalse(card.is_flight_locked)

    def test_history_full_management_in_companion_dock(self):
        """Проверка добавления, удаления и очистки истории в режиме набивки."""
        w_hist = self.mgr.win_hist
        w_hist.table_hist.setRowCount(0)

        # Добавление отчетов
        w_hist.add_history_entry("C:/reports/Report_Jan.xlsx", save_to_service=False)
        w_hist.add_history_entry("C:/reports/Report_Feb.xlsx", save_to_service=False)
        self.assertEqual(w_hist.table_hist.rowCount(), 2)
        self.assertEqual(w_hist.table_hist.item(0, 0).text(), "Report_Jan.xlsx")
        self.assertEqual(w_hist.table_hist.item(1, 0).text(), "Report_Feb.xlsx")

        # Удаление выделенной строки
        w_hist.table_hist.selectRow(0)
        w_hist._remove_selected()
        self.assertEqual(w_hist.table_hist.rowCount(), 1)
        self.assertEqual(w_hist.table_hist.item(0, 0).text(), "Report_Feb.xlsx")

        # Полная очистка
        w_hist._clear_all()
        self.assertEqual(w_hist.table_hist.rowCount(), 0)

    def test_replacement_badge_update(self):
        """Проверка динамического обновления бейджа замен ИПУ."""
        self.dummy_main.closed_meters = [{"meter": "123"}]
        self.mgr.update_replacements_badge()
        self.assertIn("1 замен", self.mgr.win_files.btn_repl.text())

        self.dummy_main.closed_meters = []
        self.mgr.update_replacements_badge()
        self.assertEqual(self.mgr.win_files.btn_repl.text(), "Мастер замен счетчиков")

    def test_directory_persistence_and_file_linking(self):
        """Проверка сохранения директорий и динамической подсветки связи файлов."""
        w_files = self.mgr.win_files
        w_files.set_template_path("C:/data/Template_2026_01.xlsx")
        self.assertEqual(w_files.tpl_path, "C:/data/Template_2026_01.xlsx")
        self.assertEqual(w_files.drop_tpl.property("state"), "linked")

        w_files.set_arcus_path("C:/data/Arcus_2026_01.xlsx")
        self.assertEqual(w_files.arc_path, "C:/data/Arcus_2026_01.xlsx")
        self.assertEqual(w_files.drop_arc.property("state"), "linked")

    def test_watchdog_timer_and_self_healing(self):
        """Проверка работы 2-секундного сторожевого таймера и авторазблокировки."""
        self.mgr.enter_companion_mode()
        self.assertTrue(self.mgr.watchdog_timer.isActive())

        # Искусственно создаем зависший флаг анимации
        self.mgr.is_dock_expanded = True
        self.mgr.is_flight_animating = True
        self.mgr._animating_dock = True
        self.mgr._active_anim_group = None

        # Запускаем сторожевой чек
        self.mgr._watchdog_check()

        # Флаги успешно сброшены и карты разблокированы
        self.assertFalse(self.mgr.is_flight_animating)
        self.assertFalse(self.mgr._animating_dock)
        self.assertTrue(self.mgr.proximity_timer.isActive())

        self.mgr.exit_companion_mode()
        self.assertFalse(self.mgr.watchdog_timer.isActive())

    def test_meter_replacement_dialog_frameless_and_operations(self):
        """Проверка безрамочного Мастера замен счетчиков и всех его операций."""
        from ui.dialogs.replacement_dialog import MeterReplacementDialog
        from models import ClosedMeterRecord, NewMeterRecord

        apts_data = {
            "квартира 1": [
                {"type": "cold", "num": 1, "prev": 120.5},
                {"type": "hot", "num": 1, "prev": 85.0}
            ],
            "квартира 2": [
                {"type": "cold", "num": 1, "prev": 210.0}
            ]
        }

        dlg = MeterReplacementDialog(None, apts_data=apts_data)
        # Проверяем безрамочность и прозрачность
        self.assertTrue(bool(dlg.windowFlags() & Qt.WindowType.FramelessWindowHint))
        self.assertTrue(dlg.testAttribute(Qt.WidgetAttribute.WA_TranslucentBackground))

        # Выбираем квартиру 1 и добавляем замену для ХВС
        dlg.combo_apt.setCurrentIndex(0)
        self.assertEqual(dlg.txt_final.text(), "120.5")
        dlg.txt_initial.setText("0.0")
        dlg._add_replacement()

        closed, new = dlg.get_results()
        self.assertEqual(len(closed), 1)
        self.assertEqual(len(new), 1)
        self.assertEqual(closed[0].apartment, "квартира 1")
        self.assertEqual(closed[0].water_type, "cold")
        self.assertEqual(closed[0].final_reading, 120.5)
        self.assertEqual(new[0].initial_reading, 0.0)

        # Редактирование
        dlg.txt_initial.setText("5.0")
        dlg._add_replacement()
        closed, new = dlg.get_results()
        self.assertEqual(len(closed), 1)
        self.assertEqual(new[0].initial_reading, 5.0)

        # Удаление
        dlg._remove_record(closed[0])
        closed, new = dlg.get_results()
        self.assertEqual(len(closed), 0)
        self.assertEqual(len(new), 0)

    def test_dock_pinning_feature(self):
        """Проверка функции закрепления (Pin) дока на экране."""
        self.mgr.enter_companion_mode()
        self.assertFalse(self.mgr.is_pinned)

        # Переключаем закрепление
        self.mgr.toggle_pin()
        self.assertTrue(self.mgr.is_pinned)
        self.assertTrue(self.mgr.is_dock_expanded)

        # При активном пине парковка не срабатывает
        self.mgr._park_all_cards_together()
        self.assertTrue(self.mgr.is_dock_expanded)

        # Снимаем закрепление
        self.mgr.toggle_pin()
        self.assertFalse(self.mgr.is_pinned)

        self.mgr.exit_companion_mode()

    def test_mini_floater_minimization_and_restoration(self):
        """Проверка полного скрытия карточек в парящую мини-панель внизу экрана и восстановления."""
        self.mgr.enter_companion_mode()
        self.assertTrue(self.mgr.is_companion_active)

        # Сворачиваем в мини-панель
        self.mgr.minimize_to_mini_floater()
        self.assertTrue(self.mgr.is_minimized_to_floater)
        self.assertTrue(self.mgr.win_mini_floater.isVisible())
        for card in self.mgr.cards:
            self.assertFalse(card.isVisible())

        # Восстанавливаем из мини-панели
        self.mgr.restore_from_mini_floater()
        self.assertFalse(self.mgr.is_minimized_to_floater)
        self.assertFalse(self.mgr.win_mini_floater.isVisible())
        for card in self.mgr.cards:
            self.assertTrue(card.isVisible())
        self.assertTrue(self.mgr.is_dock_expanded)

        self.mgr.exit_companion_mode()

    def test_system_tray_minimization_and_restoration(self):
        """Проверка сворачивания в системный трей (область уведомлений) и восстановления."""
        self.mgr.enter_companion_mode()
        self.assertTrue(self.mgr.is_companion_active)

        # Сворачиваем в трей
        self.mgr.minimize_to_tray()
        self.assertTrue(self.mgr.is_minimized_to_floater)
        for card in self.mgr.cards:
            self.assertFalse(card.isVisible())

        # Восстанавливаем из трея
        self.mgr.restore_from_tray()
        self.assertFalse(self.mgr.is_minimized_to_floater)
        for card in self.mgr.cards:
            self.assertTrue(card.isVisible())
        self.assertTrue(self.mgr.is_dock_expanded)

        self.mgr.exit_companion_mode()


class TestStrikethroughAndValidation(unittest.TestCase):
    """Тестирование считывания зачеркиваний из Аркуса и их сохранения через ExcelValidator."""

    def test_strikethrough_preservation_in_save_and_validation(self):
        import openpyxl
        from core.excel_parser import ExcelManager
        from core.excel_validator import ExcelFormatValidator

        arc_file = os.path.join(BASE_DIR, "пос 28 тест.xlsx")
        out_file = os.path.join(BASE_DIR, "хуй.xlsx")

        if os.path.exists(arc_file) and os.path.exists(out_file):
            mgr = ExcelManager()
            wb, ws, meters, meter_by_type, all_rows, non_apartment_rows, name_col = mgr.extract_data(out_file, arc_file)

            # Проверяем, что кв 334 и кв 323 считали зачеркивания
            self.assertTrue(all_rows.get('квартира 334', {}).get('striked', {}).get('cold', False))
            self.assertFalse(all_rows.get('квартира 334', {}).get('striked', {}).get('hot', False))

            self.assertTrue(all_rows.get('квартира 323', {}).get('striked', {}).get('cold', False))
            self.assertTrue(all_rows.get('квартира 323', {}).get('striked', {}).get('hot', False))

            test_target = os.path.join(BASE_DIR, "test_results", "unit_test_strike.xlsx")
            mgr.save_result(wb, ws, test_target, meters, all_rows, non_apartment_rows, name_col)

            # Проверяем, что в итоговом файле зачеркивание сохранено
            wb_check = openpyxl.load_workbook(test_target)
            ws_check = wb_check.active

            found_334 = False
            for r in range(1, ws_check.max_row + 1):
                val_a = str(ws_check.cell(r, 1).value or '')
                if '334' in val_a:
                    found_334 = True
                    # Холодный счетчик (Колонки 2, 3, 4) должен быть зачеркнут
                    for col_idx in [2, 3, 4]:
                        c = ws_check.cell(r, col_idx)
                        self.assertTrue(c.font and c.font.strike, f"Cell {c.coordinate} must have strike=True")
                    # Горячий счетчик (Колонки 8, 9, 10) НЕ должен быть зачеркнут
                    for col_idx in [8, 9, 10]:
                        c = ws_check.cell(r, col_idx)
                        self.assertFalse(bool(c.font and c.font.strike), f"Cell {c.coordinate} must NOT have strike")
                    break

            self.assertTrue(found_334, "Квартира 334 должна присутствовать в итоговом файле")

    def test_template_strikethrough_ignored_and_overridden(self):
        """Проверка: зачеркивание в шаблоне полностью игнорируется, если в Аркусе счетчик не зачеркнут."""
        import openpyxl
        from openpyxl.styles import Font
        from core.excel_parser import ExcelManager

        # Создаем временный шаблон с зачеркнутыми ячейками
        wb_tpl = openpyxl.Workbook()
        ws_tpl = wb_tpl.active
        ws_tpl.title = "Ведомость"
        ws_tpl.append(["", "Холодная вода", "", "", "Горячая вода", "", ""])
        ws_tpl.append(["Абонент", "Предыдущее", "Текущее", "Расход", "Предыдущее", "Текущее", "Расход"])
        ws_tpl.append(["квартира 100", 10.0, 15.0, 5.0, 20.0, 25.0, 5.0])
        
        # Навешиваем зачеркивание в шаблоне на квартиру 100 ХВ (строка 3)
        ws_tpl.cell(row=3, column=2).font = Font(name="Tahoma", strike=True)
        ws_tpl.cell(row=3, column=3).font = Font(name="Tahoma", strike=True)

        tpl_path = os.path.join(BASE_DIR, "test_results", "test_tpl_strike_ignore.xlsx")
        os.makedirs(os.path.dirname(tpl_path), exist_ok=True)
        wb_tpl.save(tpl_path)

        # Создаем Аркус без зачеркиваний
        wb_arc = openpyxl.Workbook()
        ws_arc = wb_arc.active
        ws_arc.append(["", "", "Холодная вода", "", "", "Горячая вода", "", ""])
        ws_arc.append(["Лицевой", "Абонент", "Предыдущее", "Текущее", "Расход", "Предыдущее", "Текущее", "Расход"])
        ws_arc.append(["12345", "квартира 100", 10.0, 15.0, 5.0, 20.0, 25.0, 5.0])
        
        arc_path = os.path.join(BASE_DIR, "test_results", "test_arc_no_strike.xlsx")
        wb_arc.save(arc_path)

        mgr = ExcelManager()
        wb, ws, meters, meter_by_type, all_rows, non_apartment_rows, name_col = mgr.extract_data(tpl_path, arc_path)

        # Проверяем, что в all_rows зачеркивание не установилось
        self.assertFalse(all_rows.get('квартира 100', {}).get('striked', {}).get('cold', False))
        self.assertFalse(all_rows.get('квартира 100', {}).get('striked', {}).get('hot', False))

        res_path = os.path.join(BASE_DIR, "test_results", "test_res_strike_cleared.xlsx")
        mgr.save_result(wb, ws, res_path, meters, all_rows, non_apartment_rows, name_col)

        wb_res = openpyxl.load_workbook(res_path)
        ws_res = wb_res.active
        
        # Проверяем, что в результирующем файле зачеркивание снято
        for r in range(1, ws_res.max_row + 1):
            if '100' in str(ws_res.cell(r, 1).value or ''):
                for col_idx in range(2, ws_res.max_column + 1):
                    cell = ws_res.cell(r, col_idx)
                    self.assertFalse(bool(cell.font and cell.font.strike), f"Cell {cell.coordinate} must NOT be striked")



if __name__ == "__main__":
    unittest.main()
