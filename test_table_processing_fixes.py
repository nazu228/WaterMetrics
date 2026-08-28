"""
test_table_processing_fixes.py — Комплексные тесты исправлений обработки таблиц:
1. Формирование названий домов по 2-й строке таблицы (а не по имени файла).
2. Чтение файлов Аркуса в любых форматах (Format A подробный и Format B сводный 1C).
3. Сопоставление Аркуса по адресу из ячеек (строка 4/содержимое).
4. Режим набивки: 100% паритет стилей во всех темах и корректное применение палитры.
5. Умное адаптивное изменение размеров ячеек и кнопок при появлении плашек.
"""

import os
import sys
import unittest

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

from PySide6.QtWidgets import QApplication

from core.excel_parser import ExcelManager
from services.folder_service import FolderNavigationService
from ui.styles import ThemeManager


class TestTableProcessingFixes(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.app = QApplication.instance() or QApplication(sys.argv)

    def test_01_house_name_from_table_row2(self):
        """Проверка: имя дома извлекается строго из 2-й строки таблицы (а не из имени файла)."""
        tpl_path = "хуй.xlsx"
        if os.path.exists(tpl_path):
            house_name = ExcelManager.extract_house_name(tpl_path)
            self.assertEqual(house_name, "Посадского 28")

            out_filename = ExcelManager.parse_house_and_next_month(tpl_path)
            self.assertEqual(out_filename, "Посадского 28.xlsx")

        tpl_path2 = "Душистая 45+.xlsx"
        if os.path.exists(tpl_path2):
            house_name2 = ExcelManager.extract_house_name(tpl_path2)
            self.assertEqual(house_name2, "Душистая 45")
            out_fn2 = ExcelManager.parse_house_and_next_month(tpl_path2)
            self.assertEqual(out_fn2, "Душистая 45.xlsx")

    def test_02_arcus_format_b_reading(self):
        """Проверка: файл Аркус формата 1С/сводного (Format B) считывается без сбоев."""
        em = ExcelManager()
        tpl = "Южный город/07 Июль 2026/3я Целиноградская 7.xlsx"
        arc = "Южный город/06 Июнь 2026/123.xlsx"

        if os.path.exists(tpl) and os.path.exists(arc):
            wb, ws, meters, meter_by_type, all_rows, non_apartment_rows, name_col = em.extract_data(tpl, arc)
            self.assertGreater(len(all_rows), 50)
            self.assertIn("квартира 1", all_rows)
            consum_1 = all_rows["квартира 1"]["consum"]
            self.assertTrue(any(v > 0 for v in consum_1.values()))

    def test_03_arcus_averkieva_format_b(self):
        """Проверка: файл Аркус Аверкиева 1234.xlsx корректно читается."""
        em = ExcelManager()
        tpl = "Южный город/07 Июль 2026/Аверкиева 34+.xlsx"
        arc = "Южный город/аркус июль/1234.xlsx"

        if os.path.exists(tpl) and os.path.exists(arc):
            wb, ws, meters, meter_by_type, all_rows, non_apartment_rows, name_col = em.extract_data(tpl, arc)
            self.assertGreater(len(all_rows), 100)
            self.assertIn("квартира 1", all_rows)
            consum_1 = all_rows["квартира 1"]["consum"]
            self.assertTrue(any(v > 0 for v in consum_1.values()))

    def test_04_arcus_content_address_matching(self):
        """Проверка: Аркус с числовым именем (123.xlsx) находится по адресу внутри ячеек строки 4."""
        addr_123 = FolderNavigationService.extract_house_name_from_arcus_content("Южный город/06 Июнь 2026/123.xlsx")
        self.assertIn("3-я Целиноградская 7", addr_123)

        matched = FolderNavigationService.is_house_match("3-я Целиноградская 7", addr_123)
        self.assertTrue(matched)

        matched_pos = FolderNavigationService.is_house_match("Посадского 28", "пос 28.xlsx")
        self.assertTrue(matched_pos)

    def test_05_companion_theme_assets_update(self):
        """Проверка: режим набивки 100% переключает и сохраняет цвета для всех тем (включая 'Как дома')."""
        from ui.main_window import MainWindow
        win = MainWindow()
        cm = win.companion_manager

        ThemeManager.apply_theme("Как дома")
        cm.update_theme_styles("Как дома")
        for card in cm.cards:
            self.assertEqual(card._current_theme_name, "Как дома")

        ThemeManager.apply_theme("Pearl Light")
        cm.update_theme_styles("Pearl Light")
        for card in cm.cards:
            self.assertEqual(card._current_theme_name, "Pearl Light")

        ThemeManager.apply_theme("Dark Tech Azure")
        cm.update_theme_styles("Dark Tech Azure")
        for card in cm.cards:
            self.assertEqual(card._current_theme_name, "Dark Tech Azure")

    def test_06_smart_navigation_ui_relayout(self):
        """Проверка: умная навигация и плашка отчета обновляют высоту карточки и релейаут дока."""
        from ui.components.companion_dock import AuthenticFilesWindow
        win_files = AuthenticFilesWindow()
        win_files.show()

        tpl_path = "Южный город/07 Июль 2026/Душистая 45.xlsx"
        if os.path.exists(tpl_path):
            win_files.set_template_path(tpl_path)
            self.assertTrue(win_files.smart_nav_frame.isVisible())
            self.assertIn("Душистая 45", win_files.lbl_smart_context.text())

    def test_07_onboarding_restart_from_any_page(self):
        """Проверка: обучение первой проводке успешно запускается из любой вкладки и окна."""
        from ui.main_window import MainWindow
        win = MainWindow()
        win.show()

        # 1. Запуск со вкладки 'О программе' (AboutPage)
        win.switch_page(4)  # AboutPage
        self.assertEqual(win.stack.currentIndex(), 4)
        win.page_about.restart_onboarding()

        self.assertIsNotNone(win.onboarding_overlay)
        self.assertTrue(win.onboarding_overlay.isVisible())
        self.assertEqual(win.stack.currentIndex(), 0)

        # 2. Навигация по шагам
        win.onboarding_overlay.next_step()
        self.assertEqual(win.onboarding_overlay.current_step_idx, 1)

        win.onboarding_overlay.prev_step()
        self.assertEqual(win.onboarding_overlay.current_step_idx, 0)

        # 3. Закрытие оверлея
        win.onboarding_overlay.skip_tour()
        self.assertFalse(win.onboarding_overlay.isVisible())

    def test_08_excel_import_dialog_and_parsing(self):
        """Проверка: интерактивный маппинг столбцов и выгрузка значений ХВС/ГВС/ДОБ."""
        from core.excel_parser import ExcelManager
        from ui.dialogs.excel_import_dialog import ExcelImportDialog

        # 1. Проверка автодетекта ролей столбцов
        sample_rows = [
            ["Адрес дома", "№ кв", "ХВС (м3)", "ГВС (м3)", "ДОБ. ХВС", "Примечание"],
            ["Душистая 45", "1", "125.4", "45.2", "12.0", "Ок"],
            ["Посадского 28", "2", "210.0", "90.5", "5.5", "Ок"],
            ["3-я Целиноградская 7", "3", "300.1", "110.0", "0.0", "Ок"]
        ]

        roles = ExcelManager.detect_column_roles(sample_rows)
        self.assertEqual(roles.get('дом'), 0)
        self.assertEqual(roles.get('хвс'), 2)
        self.assertEqual(roles.get('гвс'), 3)
        self.assertEqual(roles.get('доб'), 4)

        # 2. Проверка извлечения значений по маппингу
        values = ExcelManager.extract_values_by_mapping(sample_rows, roles, "Душистая 45")
        self.assertIsNotNone(values)
        self.assertEqual(values.get('хвс'), 125.4)
        self.assertEqual(values.get('гвс'), 45.2)
        self.assertEqual(values.get('доб'), 12.0)

        # 3. Проверка инициализации диалога
        dlg = ExcelImportDialog(None, house_name="Душистая 45")
        dlg.rows_data = sample_rows
        dlg.mapping = {0: 'house', 2: 'hvs', 3: 'gvs', 4: 'dob'}
        dlg._update_table()
        dlg._search_house_values()

        self.assertTrue(dlg.btn_apply.isEnabled())
        self.assertIn("125.4", dlg.lbl_results.text())
        dlg.close()

    def test_09_session_import_cache_and_ghost_suggestions(self):
        """Проверка: сохранение данных импорта в сессионном кэше и подстановка серых подсказок с подтверждением Enter."""
        from services.import_session_service import ImportSessionService
        from ui.dashboard_page import SmartNumericLineEdit
        from PySide6.QtGui import QKeyEvent
        from PySide6.QtCore import Qt, QEvent

        sample_rows = [
            ["Адрес дома", "№ кв", "ХВС (м3)", "ГВС (м3)", "ДОБ. ХВС"],
            ["Душистая 45", "1", "125.4", "45.2", "12.0"],
            ["2-я Целиноградская 1", "2", "599.832", "225.935", "-124.0"],
            ["Посадского 28", "3", "2144.402", "828.865", "-252.0"]
        ]
        mapping = {'дом': 0, 'хвс': 2, 'гвс': 3, 'доб': 4}

        # 1. Сохранение в кэш сессии
        ImportSessionService.set_session_import("test_import.xlsx", "Sheet1", sample_rows, mapping)
        self.assertTrue(ImportSessionService.has_active_session())

        # 2. Получение значений для дома
        vals = ImportSessionService.get_values_for_house("2-я Целиноградская 1")
        self.assertIsNotNone(vals)
        self.assertEqual(vals.get('хвс'), 599.832)
        self.assertEqual(vals.get('гвс'), 225.935)
        self.assertEqual(vals.get('доб'), -124.0)

        # 3. Проверка SmartNumericLineEdit ghost/suggested механизма
        field_c = SmartNumericLineEdit("0.0")
        field_h = SmartNumericLineEdit("0.0")
        field_d = SmartNumericLineEdit("0.0")
        chain = [field_c, field_h, field_d]
        for f in chain:
            f.linked_fields = chain

        field_c.set_suggested_value("599.832")
        field_h.set_suggested_value("225.935")
        field_d.set_suggested_value("-124.0")

        self.assertTrue(field_c.is_suggested)
        self.assertTrue(field_c.property("suggested"))
        self.assertEqual(field_c.text(), "599.832")

        # 4. Нажатие Enter подтверждает подсказку
        enter_event = QKeyEvent(QEvent.Type.KeyPress, Qt.Key.Key_Return, Qt.KeyboardModifier.NoModifier)
        field_c.keyPressEvent(enter_event)

        self.assertFalse(field_c.is_suggested)
        self.assertFalse(field_c.property("suggested"))
        self.assertFalse(field_h.is_suggested)
        self.assertFalse(field_d.is_suggested)

        ImportSessionService.clear_session()
        self.assertFalse(ImportSessionService.has_active_session())

    def test_10_robust_house_name_matching(self):
        """Проверка: надежное сопоставление сложных наименований домов (Гассия 6а, латиница, слитно, суффиксы)."""
        test_pairs = [
            ('Гассия 6а', 'Гассия, 6А', True),
            ('Гассия 6а', 'Гассия 6A', True),  # Latin 'A'
            ('Гассия 6а', 'ул. им. Валерия Гассия, 6А', True),
            ('Гассия 6а', 'Гассия 6-А', True),
            ('Гассия 6а', 'Гассия 6 А', True),
            ('Гассия 6а', 'Гассия 6/А', True),
            ('Гассия 6а', 'Гассия6а', True),
            ('Гассия 6а', 'Гассия 6', True),
            ('Гассия 6', 'Гассия, 6А', True),
            ('Гассия 6а', 'Гассия 6Б', False),  # Разные буквы
            ('Гассия 6А ЮД', 'Гассия, 6А', True),
            ('Посадского32 май 26.xlsx', 'Посадского, 32', True),
            ('3я Целиноградская 7 май 26.xlsx', '3-я Целиноградская,7', True),
            ('2я Целиноградская 11 май 26.xlsx', '2-я Целиноградская, 11', True),
            ('Дубравная 13 май 26.xlsx', 'Дубравная, 13', True),
            ('Душистая 45 май 26.xlsx', 'Душистая, 45', True),
            ('Аверкиева 34+.xlsx', 'Аверкиева, 34', True),
            ('2-я Целиноградская 1', '1-я Целиноградская, 1', True),
            ('Посадского 42 ЮД', 'Посадского,42', True),
            ('Посадского 28 май 26.xlsx', 'ул. им. Героя Владислава Посадского, 28', True),
            ('Аверкиева 38 ЮД.xlsx', 'ул. им. Героя Аверкиева А.А., 38', True),
        ]

        for n1, n2, expected in test_pairs:
            res = FolderNavigationService.is_house_match(n1, n2)
            self.assertEqual(res, expected, f"Ошибка сопоставления: {n1!r} vs {n2!r} (получено {res}, ожидалось {expected})")

    def test_11_arcus_new_cells_and_rows_red_highlighting(self):
        """Проверка: новые ячейки счетчиков (которых не было в шаблоне) и новые строки выделяются красным шрифтом."""
        import openpyxl
        from core.excel_validator import ExcelFormatValidator

        # 1. Шаблон: Квартира 1 имеет только ХВС 1 и ГВС 1 (ХВС 2 пустой: None).
        # Есть служебный блок 'Закрытые ИПУ', строка 'Итого' и подпись.
        wb_tpl = openpyxl.Workbook()
        ws_tpl = wb_tpl.active
        ws_tpl.title = "07.2026"
        ws_tpl.append(["РЕЕСТР ПОКАЗАНИЙ ЗА ИЮЛЬ 2026 ГОДА", "", "", "", "", "", "", "", "", ""])
        ws_tpl.append(["Дом Тест Ред Ячейки", "", "", "", "", "", "", "", "", ""])
        ws_tpl.append(["", "Холодная вода", "", "", "Холодная вода", "", "", "Горячая вода (ГВС)", "", ""])
        ws_tpl.append(["", "№1", "", "", "№2", "", "", "№1", "", ""])
        ws_tpl.append(["Лицевой", "Предыдущее", "Текущее", "Расход", "Предыдущее", "Текущее", "Расход", "Предыдущее", "Текущее", "Расход"])
        # Квартира 1: ХВС 1 есть, ХВС 2 пустой (None), ГВС 1 есть
        ws_tpl.append(["Квартира 1", 10.0, 15.0, 5.0, None, None, None, 20.0, 25.0, 5.0])
        ws_tpl.append(["Закрытые ИПУ", None, None, None, None, None, None, None, None, None])
        ws_tpl.append(["Квартира 88", 5.0, 5.0, 0.0, None, None, None, None, None, None])
        ws_tpl.append(["Итого", None, None, 5.0, None, None, 0.0, None, None, 5.0])
        ws_tpl.append([None] * 10)
        ws_tpl.append(["Директор ООО 'Южный город' Бочарова В.М.               ____________________"] + [None] * 9)

        test_dir = os.path.join(os.path.dirname(__file__), "test_results")
        os.makedirs(test_dir, exist_ok=True)
        tpl_path = os.path.join(test_dir, "test_tpl_new_cells.xlsx")
        wb_tpl.save(tpl_path)

        # 2. Аркус:
        # - Квартира 1: ХВС 1 штатно, ХВС 2 — НОВЫЙ СЧЕТЧИК (появились показания: 0.0, 3.0, 3.0), ГВС 1 штатно
        # - Квартира 2: НОВАЯ КВАРТИРА (отсутствовала в шаблоне)
        wb_arc = openpyxl.Workbook()
        ws_arc = wb_arc.active
        ws_arc.append([None, None, None, "Холодная вода", None, None, "Холодная вода", None, None, "Горячая вода (ГВС)", None, None])
        ws_arc.append([None, None, None, "№1", None, None, "№2", None, None, "№1", None, None])
        ws_arc.append([None, None, None, "Предыдущее", "Текущее", "Расход", "Предыдущее", "Текущее", "Расход", "Предыдущее", "Текущее", "Расход"])
        ws_arc.append(["Квартира 1", "111", "Иванов", 15.0, 20.0, 5.0, 0.0, 3.0, 3.0, 25.0, 30.0, 5.0])
        ws_arc.append(["кв. 2", "222", "Петров", 50.0, 55.0, 5.0, None, None, None, 60.0, 65.0, 5.0])

        arc_path = os.path.join(test_dir, "test_arc_new_cells.xlsx")
        wb_arc.save(arc_path)

        em = ExcelManager()
        wb, ws, meters, meter_by_type, all_rows, non_apartment_rows, name_col = em.extract_data(tpl_path, arc_path)

        self.assertIn("квартира 1", all_rows)
        self.assertIn("кв. 2", all_rows)
        # Проверяем, что новый счетчик ХВС 2 зафиксирован для Квартиры 1
        self.assertIn(('cold', 2), all_rows["квартира 1"].get("new_meter_keys", set()))
        # Проверяем, что Квартира 2 определена как новая
        self.assertTrue(all_rows["кв. 2"].get("is_new_from_arcus", False))

        res_path = os.path.join(test_dir, "test_res_new_cells.xlsx")
        em.save_result(wb, ws, res_path, meters, all_rows, non_apartment_rows, name_col)

        # 3. Проверяем цвета шрифтов в сохраненном файле
        wb_res = openpyxl.load_workbook(res_path)
        ws_res = wb_res.active

        row_apt1, row_apt2 = None, None
        for r in range(1, ws_res.max_row + 1):
            v = str(ws_res.cell(r, 1).value or '')
            if 'квартира 1' in v.lower():
                row_apt1 = r
            elif 'кв. 2' in v.lower() or 'квартира 2' in v.lower():
                row_apt2 = r

        self.assertIsNotNone(row_apt1, "Квартира 1 должна быть в таблице")
        self.assertIsNotNone(row_apt2, "Квартира 2 должна быть в таблице")

        # Квартира 1:
        # Имя (колонка 1) — обычный шрифт
        cell_name1 = ws_res.cell(row_apt1, 1)
        c_rgb = str(getattr(cell_name1.font.color, 'rgb', '')) if cell_name1.font and cell_name1.font.color else ''
        self.assertNotIn(c_rgb.upper(), ("FFFF0000", "FF0000", "00FF0000", "RED"))

        # ХВС 1 (колонки 2, 3, 4) — обычный шрифт (не красный)
        for c in (2, 3, 4):
            cell = ws_res.cell(row_apt1, c)
            c_rgb = str(getattr(cell.font.color, 'rgb', '')) if cell.font and cell.font.color else ''
            self.assertNotIn(c_rgb.upper(), ("FFFF0000", "FF0000", "00FF0000", "RED"))

        # ХВС 2 (колонки 5, 6, 7) — НОВЫЙ СЧЕТЧИК -> КРАСНЫЙ ШРИФТ!
        for c in (5, 6, 7):
            cell = ws_res.cell(row_apt1, c)
            self.assertIsNotNone(cell.font, f"Ячейка {cell.coordinate} должна иметь шрифт")
            self.assertIsNotNone(cell.font.color, f"Ячейка {cell.coordinate} должна иметь цвет шрифта")
            c_rgb = str(getattr(cell.font.color, 'rgb', '')).upper()
            self.assertIn(c_rgb, ("FFFF0000", "FF0000", "00FF0000", "RED"), f"Ячейка {cell.coordinate} должна быть красной, получено {c_rgb}")

        # ГВС 1 (колонки 8, 9, 10) — обычный шрифт
        for c in (8, 9, 10):
            cell = ws_res.cell(row_apt1, c)
            c_rgb = str(getattr(cell.font.color, 'rgb', '')) if cell.font and cell.font.color else ''
            self.assertNotIn(c_rgb.upper(), ("FFFF0000", "FF0000", "00FF0000", "RED"))

        # Квартира 2 (новая строка): ВСЕ ячейки красные!
        for c in range(1, 11):
            cell = ws_res.cell(row_apt2, c)
            if cell.value is not None:
                self.assertIsNotNone(cell.font, f"Ячейка {cell.coordinate} должна иметь шрифт")
                self.assertIsNotNone(cell.font.color, f"Ячейка {cell.coordinate} должна иметь цвет шрифта")
                c_rgb = str(getattr(cell.font.color, 'rgb', '')).upper()
                self.assertIn(c_rgb, ("FFFF0000", "FF0000", "00FF0000", "RED"), f"Ячейка {cell.coordinate} должна быть красной, получено {c_rgb}")

        # Проверяем отчет валидатора
        report = ExcelFormatValidator.validate_file(res_path, template_path=tpl_path)
        new_row_issues = [i for i in report.issues if i.category == "new_row"]
        new_meter_issues = [i for i in report.issues if i.category == "new_meter"]
        self.assertGreaterEqual(len(new_row_issues), 1)
        self.assertGreaterEqual(len(new_meter_issues), 1)

    def test_12_skip_empty_arcus_premises_without_water(self):
        """Проверка: строки из Аркуса без показаний воды (нежилые помещения/цоколи) не выгружаются в таблицу расчетов."""
        import openpyxl

        wb_tpl = openpyxl.Workbook()
        ws_tpl = wb_tpl.active
        ws_tpl.title = "07.2026"
        ws_tpl.append(["РЕЕСТР ПОКАЗАНИЙ ЗА ИЮЛЬ 2026 ГОДА", "", "", "", "", "", "", "", "", ""])
        ws_tpl.append(["Дом Тест Помещения", "", "", "", "", "", "", "", "", ""])
        ws_tpl.append(["", "Холодная вода", "", "", "Горячая вода (ГВС)", "", ""])
        ws_tpl.append(["", "№1", "", "", "№1", "", ""])
        ws_tpl.append(["Лицевой", "Предыдущее", "Текущее", "Расход", "Предыдущее", "Текущее", "Расход"])
        ws_tpl.append(["Квартира 1", 10.0, 15.0, 5.0, 20.0, 25.0, 5.0])
        ws_tpl.append(["Помещение 10", 100.0, 102.0, 2.0, 50.0, 51.0, 1.0])
        ws_tpl.append(["Итого", None, None, 7.0, None, None, 6.0])

        test_dir = os.path.join(os.path.dirname(__file__), "test_results")
        os.makedirs(test_dir, exist_ok=True)
        tpl_path = os.path.join(test_dir, "test_tpl_skip_empty.xlsx")
        wb_tpl.save(tpl_path)

        wb_arc = openpyxl.Workbook()
        ws_arc = wb_arc.active
        ws_arc.append([None, None, None, "Холодная вода", None, None, "Горячая вода (ГВС)", None, None, "Электричество", None, None])
        ws_arc.append([None, None, None, "№1", None, None, "№1", None, None, "День", None, None])
        ws_arc.append([None, None, None, "Предыдущее", "Текущее", "Расход", "Предыдущее", "Текущее", "Расход", "Предыдущее", "Текущее", "Расход"])
        # Квартира 1 (в шаблоне)
        ws_arc.append(["Квартира 1", "111", "Иванов", 15.0, 20.0, 5.0, 25.0, 30.0, 5.0, 100, 110, 10])
        # Помещение 10 (в шаблоне, но в этом месяце в Аркусе расход 0)
        ws_arc.append(["Помещение 10", "110", "Офис 10", 102.0, 102.0, 0.0, 51.0, 51.0, 0.0, None, None, None])
        # Помещение 58 (НЕТ В ШАБЛОНЕ, все колонки воды пустые None, есть только электричество!)
        ws_arc.append(["Помещение 58", "580", "Магазин 58", None, None, None, None, None, None, 300, 350, 50])
        # Помещение 59 (НЕТ В ШАБЛОНЕ, все колонки воды 0.0)
        ws_arc.append(["Помещение 59", "590", "Склад 59", 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 400, 420, 20])
        # Квартира 2 (НЕТ В ШАБЛОНЕ, НО ЕСТЬ ВОДА!)
        ws_arc.append(["Квартира 2", "222", "Петров", 50.0, 55.0, 5.0, 60.0, 65.0, 5.0, None, None, None])

        arc_path = os.path.join(test_dir, "test_arc_skip_empty.xlsx")
        wb_arc.save(arc_path)

        em = ExcelManager()
        wb, ws, meters, meter_by_type, all_rows, non_apartment_rows, name_col = em.extract_data(tpl_path, arc_path)

        # 1. Проверяем, что Квартира 1 и Помещение 10 загружены (были в шаблоне)
        self.assertIn("квартира 1", all_rows)
        self.assertIn("помещение 10", all_rows)

        # 2. Проверяем, что Помещение 58 и 59 (без воды) НЕ попали в таблицу
        self.assertNotIn("помещение 58", all_rows)
        self.assertNotIn("помещение 59", all_rows)

        # 3. Проверяем, что Квартира 2 (новая, с реальной водой) УСПЕШНО добавлена
        self.assertIn("квартира 2", all_rows)
        self.assertTrue(all_rows["квартира 2"].get("is_new_from_arcus"))


if __name__ == "__main__":
    unittest.main()


