"""
Модуль работы с Excel (openpyxl) для WaterMetrics.
Парсинг, динамическое обновление заголовков (+1 месяц) с корректным регистром ("ЗА ИЮЛЬ 2026 ГОДА"),
корректный сдвиг показаний (Новые_Предыдущие = Старые_Текущие),
профессиональное форматирование итоговой таблицы, 100% соответствие стилям исходника (Tahoma 8.5 pt, A...:M... подпись).
"""
import os
import re
from copy import copy
from datetime import datetime
import openpyxl
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

MONTHS_RU = ["январь", "февраль", "март", "апрель", "май", "июнь", "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь"]
MONTHS_RU_GEN = ["января", "февраля", "марта", "апреля", "мая", "июня", "июля", "августа", "сентября", "октября", "ноября", "декабря"]
MONTHS_RU_UPPER = ["ЯНВАРЬ", "ФЕВРАЛЬ", "МАРТ", "АПРЕЛЬ", "МАЙ", "ИЮНЬ", "ИЮЛЬ", "АВГУСТ", "СЕНТЯБРЬ", "ОКТЯБРЬ", "НОЯБРЬ", "ДЕКАБРЬ"]


class ExcelManager:
    @staticmethod
    def normalize_name(name):
        return re.sub(r'\s+', ' ', str(name or '').strip().lower())

    @staticmethod
    def canonical_apartment_key(name):
        """
        Канонический ключ квартиры/помещения для безошибочного сопоставления между шаблоном и Аркусом.
        Примеры:
        'квартира 12', 'кв. 12', 'кв 12', '12', 'Кв. 12 ' -> 'apt:12'
        'квартира 15а', 'кв. 15-а', '15А' -> 'apt:15а'
        'помещение 3', 'пом. 3', 'пом 3' -> 'pom:3'
        'офис 2' -> 'off:2'
        """
        if not name:
            return ""
        s = str(name).strip().lower().replace('ё', 'е')
        s = re.sub(r'\s+', ' ', s)
        if re.match(r'^\d+$', s):
            return f"apt:{s}"
        m_apt = re.match(r'^(?:кв\.?|квартира)\s*([\d\w\-\/]+)$', s)
        if m_apt:
            clean_num = m_apt.group(1).replace(' ', '').replace('-', '')
            return f"apt:{clean_num}"
        m_pom = re.match(r'^(?:пом\.?|помещение)\s*([\d\w\-\/]+)$', s)
        if m_pom:
            clean_num = m_pom.group(1).replace(' ', '').replace('-', '')
            return f"pom:{clean_num}"
        m_off = re.match(r'^(?:офис)\s*([\d\w\-\/]+)$', s)
        if m_off:
            clean_num = m_off.group(1).replace(' ', '').replace('-', '')
            return f"off:{clean_num}"
        return s

    @staticmethod
    def parse_float_safe(val, default=None):
        """
        Безопасное преобразование значения ячейки (int, float, str с запятой/точкой, None) в float.
        Предотвращает TypeError / ValueError при работе со свежими/пустыми шаблонами.
        """
        if val is None:
            return default
        s = str(val).strip().replace(',', '.')
        if not s:
            return default
        try:
            return round(float(s), 3)
        except (ValueError, TypeError):
            return default

    @staticmethod
    def clone_cell_style(src_cell, dst_cell):
        """Копирование всех стилей ячейки из шаблона."""
        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.border = copy(src_cell.border)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.number_format = copy(src_cell.number_format)
            dst_cell.alignment = copy(src_cell.alignment)

    @staticmethod
    def extract_house_name(wb_or_path):
        """
        Извлекает точное наименование дома из содержимого шаблона.
        По умолчанию ищет во 2-й строке (или среди первых строк таблицы), исключая служебные заголовки.
        """
        try:
            if isinstance(wb_or_path, str):
                if not os.path.exists(wb_or_path) or os.path.isdir(wb_or_path):
                    return ""
                wb = openpyxl.load_workbook(wb_or_path, data_only=True)
                ws = wb.active
            else:
                ws = wb_or_path.active if hasattr(wb_or_path, 'active') else wb_or_path

            stop_keywords = [
                'реестр', 'показаний', 'лицевого', 'коммунальные', 'предыдущее',
                'текущее', 'расход', 'холодная', 'горячая', 'квартира',
                'собственник', 'лицевой', 'содержание', '№'
            ]

            candidates = []
            for r in range(1, 7):
                for c in range(1, 10):
                    val = ws.cell(row=r, column=c).value
                    if not val:
                        continue
                    s = str(val).strip()
                    if re.match(r'^\d{4}-\d{2}-\d{2}', s):
                        continue
                    s_low = s.lower()
                    if any(k in s_low for k in stop_keywords):
                        continue

                    clean = re.sub(r'^(?:г\.\s*Краснодар,?\s*)?(?:ул\.(?:\s*им\.\s*Героя)?\s*)?', '', s, flags=re.IGNORECASE)
                    clean = re.sub(r'^(?:по\s+дому|по\s+адресу|объект|дом|ул\.)\s*:?\s*', '', clean, flags=re.IGNORECASE)
                    clean = re.sub(r'[\,\.]+$', '', clean).strip()
                    clean = re.sub(r',?\s*д\.\s*', ' ', clean).strip()

                    if re.search(r'\d', clean) and re.search(r'[А-Яа-яA-Za-z]', clean):
                        return clean
                    elif len(clean) >= 3 and not candidates:
                        candidates.append(clean)

            if candidates:
                return candidates[0]
        except Exception:
            pass

        if isinstance(wb_or_path, str):
            base = os.path.basename(wb_or_path)
            return base.replace('.xlsx', '').replace('.xls', '').replace('+', '').strip()
        return "Отчет"

    @staticmethod
    def parse_house_and_next_month(template_path):
        """Извлекает имя объекта из 2-й строки шаблона и формирует название итогового файла."""
        try:
            house_str = ExcelManager.extract_house_name(template_path)
            if not house_str:
                house_str = "Отчет"

            row2_clean = re.sub(r'\s+', '', house_str.lower())
            special_keys = ["гассия6а", "аверкиева38", "посадского42"]
            is_special = any(k in row2_clean for k in special_keys)

            if is_special:
                if not house_str.lower().endswith(" юд"):
                    final_name = f"{house_str} ЮД"
                else:
                    final_name = house_str
            else:
                final_name = house_str

            return f"{final_name}.xlsx"
        except Exception:
            return "Готово.xlsx"

    def _update_header_and_sheet_name(self, wb, ws):
        """Автоматическая смена месяца в имени листа, 1-й строке (Именительный падеж) и 2-й строке."""
        sheet_title = ws.title
        m_sheet = re.search(r'(\d{2})\.(\d{4})', sheet_title)

        if m_sheet:
            m, y = int(m_sheet.group(1)), int(m_sheet.group(2))
            next_m = 1 if m == 12 else m + 1
            next_y = y + 1 if m == 12 else y
        else:
            now = datetime.now()
            next_m = now.month
            next_y = now.year

        ws.title = f"{next_m:02d}.{next_y}"
        month_nom_upper = MONTHS_RU_UPPER[next_m - 1]

        # 1. Заголовок (1-я строка) — Именительный падеж ("ЗА ИЮЛЬ 2026 ГОДА")
        row1_cell = None
        for col in range(1, 15):
            val = ws.cell(row=1, column=col).value
            if val and str(val).strip():
                row1_cell = ws.cell(row=1, column=col)
                break

        if row1_cell:
            val_str = str(row1_cell.value)
            if re.search(r'ЗА\s+.*?\s+ГОДА', val_str, re.IGNORECASE):
                row1_cell.value = re.sub(
                    r'ЗА\s+.*?\s+ГОДА',
                    f'ЗА {month_nom_upper} {next_y} ГОДА',
                    val_str,
                    flags=re.IGNORECASE
                )
            elif re.search(r'ЗА\s+\[МЕСЯЦ\]\s+\[ГОД\]', val_str, re.IGNORECASE):
                row1_cell.value = re.sub(
                    r'ЗА\s+\[МЕСЯЦ\]\s+\[ГОД\]',
                    f'ЗА {month_nom_upper} {next_y}',
                    val_str,
                    flags=re.IGNORECASE
                )
            else:
                row1_cell.value = f"РЕЕСТР ТЕКУЩИХ И ПРЕДЫДУЩИХ ПОКАЗАНИЙ ИПУ ХОЛОДНОЙ И ГОРЯЧЕЙ ВОДЫ ЗА {month_nom_upper} {next_y} ГОДА В РАЗРЕЗЕ КАЖДОГО ЛИЦЕВОГО СЧЕТА"

        # 2. Подзаголовок (2-я строка)
        if m_sheet:
            row2_cell = None
            for col in range(1, 10):
                if ws.cell(row=2, column=col).value:
                    row2_cell = ws.cell(row=2, column=col)
                    break

            if row2_cell and row2_cell.value:
                val_str = str(row2_cell.value)
                old_m_idx = m - 1
                old_m_name = MONTHS_RU[old_m_idx]
                old_m_gen = MONTHS_RU_GEN[old_m_idx]
                new_m_name = MONTHS_RU[next_m - 1]
                new_m_gen = MONTHS_RU_GEN[next_m - 1]

                val_str = re.sub(old_m_name, new_m_name, val_str, flags=re.IGNORECASE)
                val_str = re.sub(old_m_gen, new_m_gen, val_str, flags=re.IGNORECASE)
                val_str = re.sub(str(y), str(next_y), val_str)
                row2_cell.value = val_str

    @staticmethod
    def _parse_meters(ws, detail_row, col_entries):
        def detect_water_type_from_text(text):
            if not text: return None
            s = str(text).lower()
            if 'гвс' in s or 'горячая' in s: return 'hot'
            if 'хвс' in s or 'холодная' in s: return 'cold'
            return None

        def get_cell_value_with_merged(ws_obj, r, c):
            val = ws_obj.cell(row=r, column=c).value
            if val is not None:
                return val
            for rng in ws_obj.merged_cells.ranges:
                if r in range(rng.min_row, rng.max_row + 1) and c in range(rng.min_col, rng.max_col + 1):
                    return ws_obj.cell(row=rng.min_row, column=rng.min_col).value
            return None

        def get_meter_info(group_cols, group_idx, search_rows=8):
            prev_c, curr_c, cons_c = group_cols
            water_type, meter_num = None, None
            start_row = max(1, detail_row - 1)
            end_row = max(1, detail_row - search_rows)

            min_col = prev_c
            max_col = cons_c

            for r in range(start_row, end_row - 1, -1):
                for c in range(min_col, max_col + 1):
                    val_str = str(get_cell_value_with_merged(ws, r, c) or '').strip()
                    if not val_str: continue
                    if water_type is None:
                        water_type = detect_water_type_from_text(val_str)
                    if meter_num is None:
                        for pat in [
                            r'(?:№|номер|счетчик|счётчик)\s*№?\s*(\d+)',
                            r'(\d+)\s*(?:№|номер|счетчик|счётчик)',
                            r'(?:холодная|горячая)\s*вода\s*№?\s*(\d+)',
                            r'(?:хвс|гвс)\s*№?\s*(\d+)',
                            r'сч[её]тчик\s*№?\s*(\d+)'
                        ]:
                            m = re.search(pat, val_str, re.IGNORECASE)
                            if m:
                                meter_num = int(m.group(1))
                                break
                if water_type is not None and meter_num is not None:
                    break

            if water_type is None:
                water_type = 'hot' if group_idx % 2 == 1 else 'cold'

            return water_type, meter_num

        groups = []
        i = 0
        while i + 2 < len(col_entries):
            if col_entries[i][1] == 'prev' and col_entries[i+1][1] == 'curr' and col_entries[i+2][1] == 'cons':
                groups.append((col_entries[i][0], col_entries[i+1][0], col_entries[i+2][0]))
                i += 3
            else:
                i += 1

        meters = []
        type_counter = {'cold': 0, 'hot': 0}
        for idx, cols in enumerate(groups):
            wtype, num = get_meter_info(cols, idx)
            if num is None:
                type_counter[wtype] += 1
                num = type_counter[wtype]
            else:
                type_counter[wtype] = max(type_counter.get(wtype, 0), num)
            meters.append({'type': wtype, 'num': num, 'cols': {'prev': cols[0], 'curr': cols[1], 'cons': cols[2]}})

        cold = sorted([m for m in meters if m['type'] == 'cold'], key=lambda m: m['cols']['prev'])
        hot = sorted([m for m in meters if m['type'] == 'hot'], key=lambda m: m['cols']['prev'])
        res = []
        for lst in (cold, hot):
            for i, m in enumerate(lst, 1):
                m['num'] = i
                res.append(m)
        return res

    def extract_apartments_and_meters(self, template_path, logger=None):
        """
        Извлечение структуры квартир и счетчиков.
        Используется сдвиг месяцев: стартовым значением ('prev') становится значение из колонки 'Текущие показания' прошлого месяца.
        """
        if logger:
            logger(f"ExcelManager.extract_apartments_and_meters: Извлечение из {template_path}", "INFO")
        try:
            wb = openpyxl.load_workbook(template_path, data_only=True)
            ws = wb.active

            detail_row = next((cell.row for row in ws.iter_rows(min_row=1, max_row=20, max_col=50)
                               for cell in row if cell.value and any(kw in str(cell.value).lower() for kw in ['предыдущее', 'текущее', 'расход'])), None)
            if not detail_row:
                if logger: logger("ExcelManager.extract_apartments_and_meters: Не найдена строка с показаниями", "ERROR")
                return {}

            col_entries = [(col, 'prev') if 'предыдущее' in str(ws.cell(row=detail_row, column=col).value or '').lower() else
                           (col, 'curr') if 'текущее' in str(ws.cell(row=detail_row, column=col).value or '').lower() else
                           (col, 'cons') for col in range(1, ws.max_column+1)
                           if any(kw in str(ws.cell(row=detail_row, column=col).value or '').lower() for kw in ['предыдущее', 'текущее', 'расход'])]

            meters = self._parse_meters(ws, detail_row, col_entries)
            name_col = next((col for col in range(1, 6) if ws.cell(row=detail_row, column=col).value is None), 1)

            result = {}
            in_closed_block = False

            for r in range(detail_row+1, ws.max_row+1):
                row_str = ' '.join(str(ws.cell(row=r, column=c).value or '') for c in range(1, 10)).strip().lower()
                name = str(ws.cell(row=r, column=name_col).value or '').strip()
                norm_name = self.normalize_name(name)

                if 'итого' in row_str or 'итого' in norm_name:
                    break

                if not name:
                    if in_closed_block:
                        in_closed_block = False
                    continue

                if re.search(r'(?:закрытые|замен[её]нные)\s*(?:сч[её]тчики|ипу)', norm_name) or re.search(r'(?:закрытые|замен[её]нные)\s*(?:сч[её]тчики|ипу)', row_str):
                    in_closed_block = True
                    continue

                if in_closed_block:
                    continue

                if any(ex in norm_name for ex in ['всего', 'подвал', 'чердак', 'лестница', 'коридор']):
                    continue

                m_list = []
                for m in meters:
                    # Сдвиг месяцев: Новые_Предыдущие_Показания = Старые_Текущие_Показания
                    curr_cell = ws.cell(row=r, column=m['cols']['curr']).value
                    prev_cell = ws.cell(row=r, column=m['cols']['prev']).value
                    val_to_use = curr_cell if (curr_cell is not None and str(curr_cell).strip() != '') else prev_cell
                    val = ExcelManager.parse_float_safe(val_to_use)
                    m_list.append({'type': m['type'], 'num': m['num'], 'prev': val})
                result[norm_name] = m_list

            if logger:
                logger(f"ExcelManager.extract_apartments_and_meters: Найдено {len(result)} квартир", "INFO")
            return result
        except Exception as e:
            if logger:
                logger(f"ExcelManager.extract_apartments_and_meters Ошибка: {e}", "ERROR")
            return {}

    @staticmethod
    def get_template_column_widths(template_path):
        """Извлекает точные ширины столбцов и параметры форматирования из XML шаблона."""
        widths = {}
        default_width = None
        try:
            import zipfile
            import xml.dom.minidom
            with zipfile.ZipFile(template_path) as z:
                sheet_names = [n for n in z.namelist() if n.startswith('xl/worksheets/sheet') and n.endswith('.xml')]
                if sheet_names:
                    xml_data = z.read(sheet_names[0])
                    dom = xml.dom.minidom.parseString(xml_data)
                    for sf in dom.getElementsByTagName('sheetFormatPr'):
                        if sf.hasAttribute('defaultColWidth'):
                            try:
                                default_width = float(sf.getAttribute('defaultColWidth'))
                            except Exception:
                                pass
                    for col in dom.getElementsByTagName('col'):
                        min_col = int(col.getAttribute('min'))
                        max_col = int(col.getAttribute('max'))
                        width = float(col.getAttribute('width'))
                        for c in range(min_col, max_col + 1):
                            widths[c] = width
        except Exception:
            pass
        return widths, default_width

    def extract_data(self, template_path, arcus_path, logger=None):
        """
        Основной метод загрузки данных.
        Сдвиг месяцев: Значения из колонки 'Текущие показания' шаблона (прошлый месяц) становятся
        базовыми показаниями ('prev') для нового месяца для КАЖДОГО счетчика.
        """
        if logger:
            logger("ExcelManager.extract_data: Начало загрузки данных", "INFO")

        self.template_column_widths, self.template_default_col_width = self.get_template_column_widths(template_path)

        wb = openpyxl.load_workbook(template_path, data_only=True)
        ws = wb.active

        # Сохранение текста оригинальной подписи из шаблона
        self.template_sig_text = None
        for r in range(ws.max_row, max(1, ws.max_row - 25), -1):
            val = str(ws.cell(row=r, column=1).value or '')
            if 'директор' in val.lower():
                self.template_sig_text = val.strip()
                break

        detail_row = next((cell.row for row in ws.iter_rows(min_row=1, max_row=20, max_col=50)
                           for cell in row if cell.value and any(kw in str(cell.value).lower() for kw in ['предыдущее', 'текущее', 'расход'])), None)
        if not detail_row:
            raise RuntimeError("Не найдена строка с показаниями в шаблоне.")

        col_entries = [(col, 'prev') if 'предыдущее' in str(ws.cell(row=detail_row, column=col).value or '').lower() else
                       (col, 'curr') if 'текущее' in str(ws.cell(row=detail_row, column=col).value or '').lower() else
                       (col, 'cons') for col in range(1, ws.max_column+1)
                       if any(kw in str(ws.cell(row=detail_row, column=col).value or '').lower() for kw in ['предыдущее', 'текущее', 'расход'])]

        meters = self._parse_meters(ws, detail_row, col_entries)
        meter_by_type = {'cold': [m for m in meters if m['type']=='cold'], 'hot': [m for m in meters if m['type']=='hot']}
        name_col = next((col for col in range(1, 6) if ws.cell(row=detail_row, column=col).value is None), 1)

        all_rows, non_apartment_rows = {}, set()
        all_rows_by_canon = {}
        in_closed_block = False

        for r in range(detail_row+1, ws.max_row+1):
            row_str = ' '.join(str(ws.cell(row=r, column=c).value or '') for c in range(1, 10)).strip().lower()
            name = str(ws.cell(row=r, column=name_col).value or '').strip()
            norm_name = self.normalize_name(name)
            canon_key = self.canonical_apartment_key(name)

            if 'итого' in row_str or 'итого' in norm_name:
                break

            if not name:
                if in_closed_block:
                    in_closed_block = False
                continue

            if re.search(r'(?:закрытые|замен[её]нные)\s*(?:сч[её]тчики|ипу)', norm_name) or re.search(r'(?:закрытые|замен[её]нные)\s*(?:сч[её]тчики|ипу)', row_str):
                in_closed_block = True
                non_apartment_rows.add(r)
                continue

            if in_closed_block:
                non_apartment_rows.add(r)
                continue

            if any(ex in norm_name for ex in ['всего', 'подвал', 'чердак', 'лестница', 'коридор']):
                non_apartment_rows.add(r)
                continue

            ap = {
                'row': r, 'name': name, 'norm_name': norm_name, 'canonical_key': canon_key,
                'prev': {}, 'consum': {},
                'orig_fact': {'cold': False, 'hot': False},
                'striked': {'cold': False, 'hot': False},
                'striked_meters': {},
                'is_empty': {},
                'has_3dec': {},
                'template_meters': set(),
                'new_meter_keys': set(),
                'is_new_from_arcus': False
            }

            # Сдвиг показаний для всех счетчиков абонента: Новые_Предыдущие = Старые_Текущие
            for m in meters:
                curr_cell_obj = ws.cell(row=r, column=m['cols']['curr'])
                prev_cell_obj = ws.cell(row=r, column=m['cols']['prev'])
                cons_cell_obj = ws.cell(row=r, column=m['cols']['cons'])
                curr_cell = curr_cell_obj.value
                prev_cell = prev_cell_obj.value
                cons_cell = cons_cell_obj.value

                has_data = any(v is not None and str(v).strip() != '' for v in (curr_cell, prev_cell, cons_cell))
                m_key = (m['type'], m['num'])
                if has_data:
                    val_to_use = curr_cell if (curr_cell is not None and str(curr_cell).strip() != '') else prev_cell
                    prev_val = ExcelManager.parse_float_safe(val_to_use)
                    ap['prev'][m_key] = prev_val
                    ap['template_meters'].add(m_key)
                else:
                    ap['prev'][m_key] = None

                ap['consum'][m_key] = 0.0

            all_rows[norm_name] = ap
            if canon_key:
                all_rows_by_canon[canon_key] = ap

        self.template_apartments = set(all_rows.keys())
        self.template_apartment_keys = set(all_rows_by_canon.keys())

        # Загрузка данных расхода из файла Аркус
        wb_arc = openpyxl.load_workbook(arcus_path, data_only=True)
        ws_arc = wb_arc.active
        arc_detail = next((cell.row for row in ws_arc.iter_rows(min_row=1, max_row=20, max_col=50)
                           for cell in row if cell.value and any(kw in str(cell.value).lower() for kw in ['предыдущее', 'текущее', 'расход'])), None)

        if arc_detail:
            # FORMAT A: Детализированный формат выгрузки Аркус (Предыдущее, Текущее, Расход)
            arc_col_entries = [(col, 'prev') if 'предыдущее' in str(ws_arc.cell(row=arc_detail, column=col).value or '').lower() else
                               (col, 'curr') if 'текущее' in str(ws_arc.cell(row=arc_detail, column=col).value or '').lower() else
                               (col, 'cons') for col in range(1, ws_arc.max_column+1)
                               if any(kw in str(ws_arc.cell(row=arc_detail, column=col).value or '').lower() for kw in ['предыдущее', 'текущее', 'расход'])]

            arc_meters = self._parse_meters(ws_arc, arc_detail, arc_col_entries)

            template_to_arcus = {}
            for t_m in meters:
                key = (t_m['type'], t_m['num'])
                template_to_arcus[key] = next((a_m['cols'] for a_m in arc_meters if a_m['type'] == key[0] and a_m['num'] == key[1]), None)

            arc_name_col = next((col for col in range(1, 6) if ws_arc.cell(row=arc_detail, column=col).value is None or any(kw in str(ws_arc.cell(row=arc_detail, column=col).value or '').lower() for kw in ['квартира', 'кв', 'помещение', 'абонент'])), 1)
            for r in range(arc_detail+1, ws_arc.max_row+1):
                raw_arc_val = ws_arc.cell(row=r, column=arc_name_col).value
                if not raw_arc_val:
                    continue
                arc_name = self.normalize_name(raw_arc_val)
                arc_canon = self.canonical_apartment_key(raw_arc_val)

                ap = all_rows.get(arc_name) or (all_rows_by_canon.get(arc_canon) if arc_canon else None)
                if not ap:
                    if any(ex in arc_name for ex in ['итого', 'всего', 'закрытые', 'замененные', 'директор', 'подпись', 'реестр']):
                        continue

                    # Проверяем, есть ли в Аркусе хотя бы по одному счетчику реальные показания воды (расход, текущие или предыдущие)
                    has_any_water_reading = False
                    for key, arc_cols in template_to_arcus.items():
                        if not arc_cols:
                            continue
                        for role in ('cons', 'curr', 'prev'):
                            if role in arc_cols:
                                c_obj = ws_arc.cell(row=r, column=arc_cols[role])
                                if c_obj is not None and c_obj.value is not None:
                                    v_str = str(c_obj.value).strip()
                                    if v_str not in ('', '-'):
                                        flt_val = ExcelManager.parse_float_safe(c_obj.value)
                                        if flt_val is not None and (flt_val > 0.0001 or flt_val < -0.0001):
                                            has_any_water_reading = True
                                            break
                        if has_any_water_reading:
                            break

                    if not has_any_water_reading:
                        # Строка из Аркуса отсутствует в шаблоне и не содержит показаний воды (например, нежилое помещение/цоколь без счетчиков воды)
                        continue

                    # Новая строка из Аркуса с показаниями воды, которой не было в шаблоне
                    next_r = max((a['row'] for a in all_rows.values()), default=5) + 1
                    ap_display_name = str(raw_arc_val).strip()
                    ap = {
                        'row': next_r, 'name': ap_display_name, 'norm_name': arc_name, 'canonical_key': arc_canon,
                        'prev': {}, 'consum': {},
                        'orig_fact': {'cold': False, 'hot': False},
                        'striked': {'cold': False, 'hot': False},
                        'striked_meters': {},
                        'is_empty': {},
                        'has_3dec': {},
                        'template_meters': set(),
                        'is_new_from_arcus': True,
                        'new_meter_keys': set()
                    }
                    for m in meters:
                        m_key = (m['type'], m['num'])
                        ap['prev'][m_key] = 0.0
                        ap['consum'][m_key] = 0.0
                    all_rows[arc_name] = ap
                    if arc_canon:
                        all_rows_by_canon[arc_canon] = ap
                    if logger:
                        logger(f"ExcelManager: Обнаружена новая строка с водой из Аркуса (отсутствовала в шаблоне): {ap_display_name}", "INFO")

                if 'new_meter_keys' not in ap:
                    ap['new_meter_keys'] = set()

                for key, arc_cols in template_to_arcus.items():
                    if arc_cols and 'cons' in arc_cols:
                        cell_cons = ws_arc.cell(row=r, column=arc_cols['cons'])
                        cell_val = cell_cons.value
                        cell_curr = ws_arc.cell(row=r, column=arc_cols['curr']) if 'curr' in arc_cols else None
                        cell_prev = ws_arc.cell(row=r, column=arc_cols['prev']) if 'prev' in arc_cols else None

                        is_empty = (cell_val is None or str(cell_val).strip() == '')
                        has_any_reading = any(c is not None and c.value is not None and str(c.value).strip() != '' for c in (cell_cons, cell_curr, cell_prev))
                        ap['is_empty'][key] = is_empty

                        cell_str = str(cell_val).strip() if cell_val is not None else ''
                        has_3_decimals = False
                        if '.' in cell_str or ',' in cell_str:
                            parts = re.split(r'[\.,]', cell_str)
                            if len(parts) == 2 and len(parts[1]) == 3:
                                has_3_decimals = True

                        val = ExcelManager.parse_float_safe(cell_val, default=0.0)
                        if abs(round(val, 2) - val) > 1e-5:
                            has_3_decimals = True

                        ap['consum'][key] = val
                        ap['has_3dec'][key] = has_3_decimals

                        if val > 0.0001:
                            ap['orig_fact'][key[0]] = True

                        # Если счетчик отсутствовал в шаблоне или это новая строка:
                        is_meter_new = ap.get('is_new_from_arcus') or (key not in ap.get('template_meters', set())) or (ap['prev'].get(key) is None)
                        if is_meter_new and has_any_reading:
                            ap['new_meter_keys'].add(key)
                            arc_prev_val = ExcelManager.parse_float_safe(cell_prev.value if cell_prev else None)
                            if arc_prev_val is not None:
                                ap['prev'][key] = arc_prev_val
                            elif cell_curr is not None and cell_curr.value is not None:
                                arc_curr_val = ExcelManager.parse_float_safe(cell_curr.value)
                                if arc_curr_val is not None:
                                    ap['prev'][key] = round(arc_curr_val - val, 3)
                            elif val > 0.0001:
                                ap['prev'][key] = 0.0

                        # Проверка зачеркивания в Аркусе по любой из ячеек счетчика (расход, текущее, предыдущее)
                        is_striked = False
                        for c_obj in (cell_cons, cell_curr, cell_prev):
                            if c_obj and c_obj.font and c_obj.font.strike:
                                is_striked = True
                                break

                        if is_striked:
                            ap['striked'][key[0]] = True
                            if 'striked_meters' not in ap:
                                ap['striked_meters'] = {}
                            ap['striked_meters'][key] = True
                    else:
                        ap['consum'][key] = 0.0
                        ap['is_empty'][key] = True
                        ap['has_3dec'][key] = False
        else:
            # FORMAT B: Сводный формат Аркус / 1С (Колонки расхода напрямую: 'Холодная вода', 'Холодная вода (ГВС)')
            hdr_row = next((cell.row for row in ws_arc.iter_rows(min_row=1, max_row=10, max_col=30)
                            for cell in row if cell.value and any(kw in str(cell.value).lower() for kw in ['холодная', 'гвс', 'горячая', 'квартира'])), 2)

            apt_col = next((col for col in range(1, 10) if any(kw in str(ws_arc.cell(row=hdr_row, column=col).value or '').lower() for kw in ['квартира', 'кв', 'помещение', 'абонент'])), 1)

            cold_cols = []
            hot_cols = []
            for col in range(1, ws_arc.max_column + 1):
                val = str(ws_arc.cell(row=hdr_row, column=col).value or '').lower()
                if any(ex in val for ex in ['перерасчет', 'содержание', 'пени', 'взнос', 'начислено', 'долг']):
                    continue
                if 'гвс' in val or 'горячая' in val:
                    hot_cols.append(col)
                elif 'холодная' in val or 'хвс' in val:
                    cold_cols.append(col)

            col_to_meter = {}
            for idx, col in enumerate(cold_cols, 1):
                col_to_meter[col] = ('cold', idx)
            for idx, col in enumerate(hot_cols, 1):
                col_to_meter[col] = ('hot', idx)

            for r in range(hdr_row + 1, ws_arc.max_row + 1):
                raw_arc_val = ws_arc.cell(row=r, column=apt_col).value
                if not raw_arc_val:
                    continue
                arc_name = self.normalize_name(raw_arc_val)
                arc_canon = self.canonical_apartment_key(raw_arc_val)

                ap = all_rows.get(arc_name) or (all_rows_by_canon.get(arc_canon) if arc_canon else None)
                if not ap:
                    if any(ex in arc_name for ex in ['итого', 'всего', 'закрытые', 'замененные', 'директор', 'подпись', 'реестр']):
                        continue

                    # Проверяем, есть ли в Аркусе хотя бы по одной колонке реальные показания расхода воды
                    has_any_water_reading = False
                    for col in col_to_meter.keys():
                        c_obj = ws_arc.cell(row=r, column=col)
                        if c_obj is not None and c_obj.value is not None:
                            v_str = str(c_obj.value).strip()
                            if v_str not in ('', '-'):
                                flt_val = ExcelManager.parse_float_safe(c_obj.value)
                                if flt_val is not None and (flt_val > 0.0001 or flt_val < -0.0001):
                                    has_any_water_reading = True
                                    break
                    if not has_any_water_reading:
                        continue

                    next_r = max((a['row'] for a in all_rows.values()), default=5) + 1
                    ap_display_name = str(raw_arc_val).strip()
                    ap = {
                        'row': next_r, 'name': ap_display_name, 'norm_name': arc_name, 'canonical_key': arc_canon,
                        'prev': {}, 'consum': {},
                        'orig_fact': {'cold': False, 'hot': False},
                        'striked': {'cold': False, 'hot': False},
                        'striked_meters': {},
                        'is_empty': {},
                        'has_3dec': {},
                        'template_meters': set(),
                        'is_new_from_arcus': True,
                        'new_meter_keys': set()
                    }
                    for m in meters:
                        m_key = (m['type'], m['num'])
                        ap['prev'][m_key] = 0.0
                        ap['consum'][m_key] = 0.0
                    all_rows[arc_name] = ap
                    if arc_canon:
                        all_rows_by_canon[arc_canon] = ap
                    if logger:
                        logger(f"ExcelManager: Обнаружена новая строка с водой из Аркуса (отсутствовала в шаблоне): {ap_display_name}", "INFO")

                if 'new_meter_keys' not in ap:
                    ap['new_meter_keys'] = set()

                for col, key in col_to_meter.items():
                    cell = ws_arc.cell(row=r, column=col)
                    cell_val = cell.value
                    is_empty = (cell_val is None or str(cell_val).strip() == '')
                    ap['is_empty'][key] = is_empty

                    val = ExcelManager.parse_float_safe(cell_val, default=0.0)
                    cell_str = str(cell_val).strip() if cell_val is not None else ''
                    has_3_decimals = False
                    if '.' in cell_str or ',' in cell_str:
                        parts = re.split(r'[\.,]', cell_str)
                        if len(parts) == 2 and len(parts[1]) == 3:
                            has_3_decimals = True
                    if abs(round(val, 2) - val) > 1e-5:
                        has_3_decimals = True

                    ap['consum'][key] = val
                    ap['has_3dec'][key] = has_3_decimals

                    if val > 0.0001:
                        ap['orig_fact'][key[0]] = True

                    is_meter_new = ap.get('is_new_from_arcus') or (key not in ap.get('template_meters', set())) or (ap['prev'].get(key) is None)
                    if is_meter_new and not is_empty:
                        ap['new_meter_keys'].add(key)
                        if ap['prev'].get(key) is None:
                            ap['prev'][key] = 0.0

                    if cell.font and cell.font.strike:
                        ap['striked'][key[0]] = True
                        if 'striked_meters' not in ap:
                            ap['striked_meters'] = {}
                        ap['striked_meters'][key] = True

        if logger:
            logger(f"ExcelManager.extract_data: Загружено квартир: {len(all_rows)}", "INFO")
        return wb, ws, meters, meter_by_type, all_rows, non_apartment_rows, name_col

    def save_result(self, wb, ws, save_path, meters, all_rows, non_apartment_rows, name_col, closed_meters=None, new_meters=None):
        """
        Запись результатов с гарантированным созданием папок, выравниванием рамок и шрифтов (100% Tahoma 8.5 pt),
        выделением новых строк и новых показаний красным шрифтом, оформлением итоговой строки, блока 'Закрытые ИПУ' и подписью.
        """
        save_dir = os.path.dirname(os.path.abspath(save_path))
        if save_dir:
            os.makedirs(save_dir, exist_ok=True)

        self._update_header_and_sheet_name(wb, ws)

        total_cols = max((m['cols']['cons'] for m in meters), default=ws.max_column)
        new_dict = {(nm.apartment, nm.water_type, nm.meter_num): nm.initial_reading for nm in (new_meters or [])}
        new_dict_canon = {(self.canonical_apartment_key(nm.apartment), nm.water_type, nm.meter_num): nm.initial_reading for nm in (new_meters or [])}

        start_data_r = min((ap['row'] for ap in all_rows.values()), default=6)

        # Безопасно снимаем ВСЕ объединенные диапазоны внутри и ниже диапазона строк квартир
        for rng in list(ws.merged_cells.ranges):
            if rng.max_row >= start_data_r:
                try:
                    ws.unmerge_cells(str(rng))
                except Exception:
                    pass
                if rng in ws.merged_cells.ranges:
                    try:
                        ws.merged_cells.ranges.remove(rng)
                    except Exception:
                        pass

        template_canon_keys = getattr(self, 'template_apartment_keys', set())

        # 1. Запись основной таблицы (каждый счетчик получает свои новые предыдущие и текущие показания)
        for idx, ap in enumerate(all_rows.values()):
            cur_r = start_data_r + idx
            ap['row'] = cur_r
            ap_canon = ap.get('canonical_key') or self.canonical_apartment_key(ap['name'])
            is_new_row = ap.get('is_new_from_arcus', False) or (template_canon_keys and ap_canon not in template_canon_keys)

            # Заполняем имя квартиры / абонента
            cell_name = ws.cell(row=cur_r, column=name_col)
            cell_name.value = ap['name']
            if is_new_row:
                cell_name.font = Font(name="Tahoma", size=8.5, color="FFFF0000")
            else:
                cell_name.font = Font(name="Tahoma", size=8.5)

            for m in meters:
                key = (m['type'], m['num'])
                is_replacement = ((ap_canon, m['type'], m['num']) in new_dict_canon) or ((ap['norm_name'], m['type'], m['num']) in new_dict)
                is_new_meter = (key in ap.get('new_meter_keys', set())) or is_replacement
                is_red = is_new_row or is_new_meter

                pval = new_dict_canon.get((ap_canon, m['type'], m['num']), new_dict.get((ap['norm_name'], m['type'], m['num']), ap['prev'].get(key)))
                if pval is not None:
                    pval = round(float(pval), 3)

                cons = round(float(ap['consum'].get(key, 0.0)), 3)

                cell_prev = ws.cell(row=cur_r, column=m['cols']['prev'])
                cell_curr = ws.cell(row=cur_r, column=m['cols']['curr'])
                cell_cons = ws.cell(row=cur_r, column=m['cols']['cons'])

                if pval is None and not is_new_meter:
                    if cons > 0.0001:
                        pval = 0.0
                        cell_prev.value = 0.0
                        curr_val = round(cons, 3)
                        cell_curr.value = curr_val
                        cell_cons.value = cons
                    else:
                        cell_prev.value = None
                        cell_curr.value = None
                        cell_cons.value = None
                        nf_plain = Font(name="Tahoma", size=8.5, strike=False, color="FFFF0000" if is_red else None)
                        cell_prev.font = nf_plain
                        cell_curr.font = nf_plain
                        cell_cons.font = nf_plain
                        continue
                else:
                    if pval is None:
                        pval = 0.0
                    curr_val = round(pval + cons, 3)
                    cell_prev.value = pval
                    cell_curr.value = curr_val
                    cell_cons.value = cons

                # Применение зачеркивания (строго из Аркуса)
                is_meter_striked = (
                    ap.get('striked_meters', {}).get(key, False) or
                    ap.get('striked', {}).get(m['type'], False)
                )

                if is_red:
                    nf = Font(name="Tahoma", size=8.5, strike=bool(is_meter_striked), color="FFFF0000")
                else:
                    nf = Font(name="Tahoma", size=8.5, strike=bool(is_meter_striked))
                cell_prev.font = copy(nf)
                cell_curr.font = copy(nf)
                cell_cons.font = copy(nf)

        last_ap_row = start_data_r + len(all_rows) - 1

        # Удаляем все старые строки ниже последней квартиры (старый блок закрытых ИПУ, старый Итого, старая подпись)
        if ws.max_row > last_ap_row:
            ws.delete_rows(last_ap_row + 1, ws.max_row - last_ap_row)

        cur_r = last_ap_row + 1

        # 3. Эталонные стили и цвета колонок напрямую из строки данных шаблона (строка 6)
        ref_row = 6 if ws.max_row >= 6 else (last_ap_row if last_ap_row > 0 else 1)
        col_fonts = {c: copy(ws.cell(row=ref_row, column=c).font) for c in range(1, total_cols + 1)}
        col_borders = {c: copy(ws.cell(row=ref_row, column=c).border) for c in range(1, total_cols + 1)}
        col_fills = {c: copy(ws.cell(row=ref_row, column=c).fill) for c in range(1, total_cols + 1)}
        col_num_fmts = {c: ws.cell(row=ref_row, column=c).number_format for c in range(1, total_cols + 1)}

        font_sig = Font(name="Tahoma", size=11.0, bold=False)

        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")

        # Вставка блока 'Закрытые ИПУ' (если они есть) с точными стилями колонок шаблона
        if closed_meters:
            ws.row_dimensions[cur_r].height = 17.25
            cell_hdr = ws.cell(row=cur_r, column=name_col)
            cell_hdr.value = "Закрытые ИПУ"
            if col_fonts.get(name_col):
                cell_hdr.font = copy(col_fonts[name_col])
            cell_hdr.alignment = align_left

            for c in range(1, total_cols + 1):
                cell = ws.cell(row=cur_r, column=c)
                if col_borders.get(c):
                    cell.border = copy(col_borders[c])
                if col_fonts.get(c):
                    cell.font = copy(col_fonts[c])
                if col_fills.get(c):
                    cell.fill = copy(col_fills[c])
                if c != name_col:
                    cell.value = None

            cur_r += 1

            grouped = {}
            for cm in closed_meters:
                grouped.setdefault(cm.apartment, []).append(cm)

            for apt_name, rec_list in grouped.items():
                ws.row_dimensions[cur_r].height = 17.25
                apt_display = apt_name.strip()
                if apt_display:
                    apt_display = apt_display[0].upper() + apt_display[1:]

                cell_name = ws.cell(row=cur_r, column=name_col)
                cell_name.value = apt_display
                if col_fonts.get(name_col):
                    cell_name.font = copy(col_fonts[name_col])
                cell_name.alignment = align_left
                if col_borders.get(name_col):
                    cell_name.border = copy(col_borders[name_col])

                for c in range(1, total_cols + 1):
                    cell = ws.cell(row=cur_r, column=c)
                    if col_fonts.get(c):
                        cell.font = copy(col_fonts[c])
                    if col_borders.get(c):
                        cell.border = copy(col_borders[c])
                    if col_fills.get(c):
                        cell.fill = copy(col_fills[c])
                    if col_num_fmts.get(c):
                        cell.number_format = col_num_fmts[c]
                    if c != name_col:
                        cell.alignment = align_right

                for cm in rec_list:
                    target_m = next((m for m in meters if m['type'] == cm.water_type and m['num'] == cm.meter_num), None)
                    if target_m:
                        fin_val = round(float(cm.final_reading), 3)
                        ws.cell(row=cur_r, column=target_m['cols']['prev']).value = fin_val
                        ws.cell(row=cur_r, column=target_m['cols']['curr']).value = fin_val
                        ws.cell(row=cur_r, column=target_m['cols']['cons']).value = 0.0

                cur_r += 1

        # 4. Оформление и пересчет сумм в единственной итоговой строке 'Итого' (1:1 стили и цвета шаблона)
        tot_r = cur_r
        ws.row_dimensions[tot_r].height = 15.0

        for c in range(1, total_cols + 1):
            cell = ws.cell(row=tot_r, column=c)
            cell.value = None
            if col_fonts.get(c):
                cell.font = copy(col_fonts[c])
            if col_borders.get(c):
                cell.border = copy(col_borders[c])
            if col_fills.get(c):
                cell.fill = copy(col_fills[c])
            if col_num_fmts.get(c):
                cell.number_format = col_num_fmts[c]
            if c != name_col:
                cell.alignment = align_right

        cell_tot_lbl = ws.cell(row=tot_r, column=name_col)
        cell_tot_lbl.value = "Итого"
        if col_fonts.get(name_col):
            cell_tot_lbl.font = copy(col_fonts[name_col])
        cell_tot_lbl.alignment = align_left

        for m in meters:
            sum_cons = round(sum(
                float(ws.cell(row=ap['row'], column=m['cols']['cons']).value or 0.0)
                for ap in all_rows.values()
            ), 3)
            ws.cell(row=tot_r, column=m['cols']['cons']).value = sum_cons

        # 5. Очистка устаревших объединенных ячеек и фантомных столбцов/строк
        for range_ in list(ws.merged_cells.ranges):
            if range_.max_row >= tot_r or range_.min_row >= tot_r:
                try:
                    ws.unmerge_cells(str(range_))
                except Exception:
                    pass
                if range_ in ws.merged_cells.ranges:
                    try:
                        ws.merged_cells.ranges.remove(range_)
                    except Exception:
                        pass

        if ws.max_column > total_cols:
            ws.delete_cols(total_cols + 1, ws.max_column - total_cols)

        # Форматирование 2-й строки: объединенная ячейка по ширине таблицы, выравнивание по центру.
        # 1-я строка не меняет форматирование.
        for range_ in list(ws.merged_cells.ranges):
            if range_.min_row == 2 and range_.max_row == 2:
                try:
                    ws.unmerge_cells(str(range_))
                except Exception:
                    pass
                if range_ in ws.merged_cells.ranges:
                    try:
                        ws.merged_cells.ranges.remove(range_)
                    except Exception:
                        pass
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
        ws.cell(row=2, column=1).alignment = align_center

        # 6. Динамическая подпись и очистка промежуточных строк
        sig_row = tot_r + 3

        # Очищаем промежуточные пустые строки от старых данных и границ
        for empty_r in range(tot_r + 1, sig_row):
            for c in range(1, total_cols + 1):
                cell = ws.cell(row=empty_r, column=c)
                cell.value = None
                cell.border = Border()
                cell.fill = PatternFill(fill_type=None)

        # Строка подписи
        for c in range(1, total_cols + 1):
            cell = ws.cell(row=sig_row, column=c)
            cell.value = None
            cell.border = Border()
            cell.fill = PatternFill(fill_type=None)

        sig_text = getattr(self, 'template_sig_text', None) or 'Директор ООО "Южный город"  Бочарова В.М.               ____________________'

        sig_cell = ws.cell(row=sig_row, column=1)
        sig_cell.value = sig_text
        sig_cell.font = font_sig
        sig_cell.alignment = Alignment(horizontal="right", vertical="center")

        ws.merge_cells(start_row=sig_row, start_column=1, end_row=sig_row, end_column=total_cols)

        # Удаление фантомных строк исходного шаблона после строки подписи
        if ws.max_row > sig_row:
            ws.delete_rows(sig_row + 1, ws.max_row - sig_row)

        # 8. Гарантированное сохранение ширины колонок в точности как в исходном шаблоне
        if hasattr(self, 'template_column_widths') and self.template_column_widths:
            for col in range(1, total_cols + 1):
                col_letter = get_column_letter(col)
                if col in self.template_column_widths:
                    ws.column_dimensions[col_letter].width = self.template_column_widths[col]
                elif self.template_default_col_width:
                    ws.column_dimensions[col_letter].width = self.template_default_col_width

        # 9. Финальная комплексная валидация и авто-исправление оформления (ExcelFormatValidator)
        try:
            from core.excel_validator import ExcelFormatValidator
            ExcelFormatValidator.auto_fix_and_validate(
                ws,
                total_cols=total_cols,
                sig_text=sig_text,
                template_widths=getattr(self, 'template_column_widths', None),
                template_apartments=getattr(self, 'template_apartments', None)
            )
        except Exception as ex:
            pass

        try:
            wb.save(save_path)
        except PermissionError:
            base, ext = os.path.splitext(save_path)
            timestamp = datetime.now().strftime("%H%M%S")
            fallback_path = f"{base}_{timestamp}{ext}"
            wb.save(fallback_path)
            raise PermissionError(
                f"Файл '{os.path.basename(save_path)}' открыт в Excel! "
                f"Закройте файл в Excel и повторите попытку. (Файл сохранен как '{os.path.basename(fallback_path)}')"
            )

    @staticmethod
    def preview_worksheet(path, sheet_name=None, max_rows=None):
        """
        Считывает строки листа Excel с автоподгонкой колонок.
        По умолчанию считывает все строки с данными (max_rows=None).
        """
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames

            if sheet_name and sheet_name in sheet_names:
                ws = wb[sheet_name]
                active_sheet_name = sheet_name
            else:
                ws = wb.active
                active_sheet_name = ws.title

            rows_data = []
            max_cols = 0

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if max_rows is not None and i >= max_rows:
                    break
                row_vals = [str(cell).strip() if cell is not None else "" for cell in row]
                if any(row_vals):
                    rows_data.append(row_vals)
                    if len(row_vals) > max_cols:
                        max_cols = len(row_vals)

            for r in rows_data:
                r.extend([""] * (max_cols - len(r)))

            wb.close()
            return sheet_names, rows_data, active_sheet_name
        except Exception:
            return [], [], ""

    @staticmethod
    def detect_column_roles(rows_data, max_header_rows=40):
        """
        Автоматически определяет роли колонок (дом, хвс, гвс, доб) по ключевым словам в шапках таблицы.
        """
        roles = {}
        keywords = {
            'дом': ['дом', 'адрес', 'объект', 'улица', 'ул.', 'название', 'мкд', 'наименование', 'здание', 'потребитель'],
            'хвс': ['хвс', 'холодн', 'хол.', 'хол ', 'х/в', 'холодная', 'cold'],
            'гвс': ['гвс', 'горяч', 'гор.', 'гор ', 'г/в', 'горячая', 'hot'],
            'доб': ['доб', 'добавк', 'добавление', 'доп.', 'корр', 'коррект', 'добавоч', 'разница', 'добавка']
        }

        for i, row in enumerate(rows_data):
            if i >= max_header_rows:
                break
            for col_idx, cell_val in enumerate(row):
                val_lower = cell_val.lower().strip()
                if not val_lower:
                    continue
                for role, kw_list in keywords.items():
                    if role not in roles:
                        if any(kw in val_lower for kw in kw_list):
                            roles[role] = col_idx
                            break
        return roles

    @staticmethod
    def extract_values_by_mapping(rows_data, mapping, house_name, header_rows=0):
        """
        Извлекает числовые показатели ХВС, ГВС и ДОБ для указанного дома.
        Находит адрес в любом положении строки, строго проверяя, чтобы значения метрик были числовыми.
        """
        if not house_name or not rows_data:
            return None

        from services.folder_service import FolderNavigationService

        house_col = mapping.get('дом')
        col_hvs = mapping.get('хвс')
        col_gvs = mapping.get('гвс')
        col_dob = mapping.get('доб')

        # Если ни ХВС ни ГВС не назначены, извлечь невозможно
        if col_hvs is None and col_gvs is None:
            return None

        for r_idx, row in enumerate(rows_data):
            if r_idx < header_rows:
                continue

            matched = False
            matched_cell = ""

            # 1. Проверяем назначенный столбец адреса
            if house_col is not None and house_col < len(row):
                cell_val = row[house_col]
                if cell_val and FolderNavigationService.is_house_match(house_name, cell_val):
                    matched = True
                    matched_cell = cell_val

            # 2. Если в назначенном столбце не сошлось или он не задан — сканируем все ячейки строки
            if not matched:
                for c_idx, cell in enumerate(row):
                    if cell and FolderNavigationService.is_house_match(house_name, cell):
                        matched = True
                        matched_cell = cell
                        break

            if matched:
                # 3. Проверка: значения в целевых колонках должны быть валидными числами
                val_hvs = ExcelManager.parse_float_safe(row[col_hvs]) if col_hvs is not None and col_hvs < len(row) else None
                val_gvs = ExcelManager.parse_float_safe(row[col_gvs]) if col_gvs is not None and col_gvs < len(row) else None
                val_dob = ExcelManager.parse_float_safe(row[col_dob]) if col_dob is not None and col_dob < len(row) else None

                # Должно присутствовать хотя бы одно реальное число (не текст, не пустая ячейка)
                if val_hvs is not None or val_gvs is not None or val_dob is not None:
                    return {
                        'хвс': val_hvs if val_hvs is not None else 0.0,
                        'гвс': val_gvs if val_gvs is not None else 0.0,
                        'доб': val_dob if val_dob is not None else 0.0,
                        'row_idx': r_idx,
                        'matched_cell': matched_cell
                    }

        return None