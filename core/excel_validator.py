"""
Модуль глубокой проверки и гарантированного авто-исправления оформления Excel (WaterMetrics).
Обеспечивает 100% соответствие эталонным стандартам:
- Шрифты (Tahoma 8.25/8.5 pt, Calibri 11 pt, Tahoma 11 pt)
- Границы (тонкие серые #808080 на всех 4-х сторонах ячеек таблицы, включая объединенные)
- Полная зачистка рамок в промежуточных строках и строке подписи
- Выравнивание (номера квартир - влево, числа/суммы - вправо, шапка - по центру)
- Форматы чисел (@ для квартир, ##########0.##### для показаний)
- Разметка страницы и печати (сквозные строки $3:$5, fitToWidth=1, ориентация листа)
- Физическое удаление фантомных строк и столбцов
"""
import os
import re
from copy import copy
import openpyxl
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class ValidationIssue:
    def __init__(self, category, location, message, severity="WARNING", fixed=False):
        self.category = category      # 'font', 'border', 'alignment', 'num_fmt', 'print', 'phantom', 'structure'
        self.location = location      # e.g. 'A6', 'Row 175', 'Sheet'
        self.message = message
        self.severity = severity      # 'INFO', 'WARNING', 'ERROR'
        self.fixed = fixed

    def __repr__(self):
        status = "[FIXED]" if self.fixed else f"[{self.severity}]"
        return f"{status} {self.category.upper()} at {self.location}: {self.message}"


class ValidationReport:
    def __init__(self, filename="Workbook"):
        self.filename = filename
        self.issues = []
        self.fixes = []
        self.stats = {}

    @property
    def is_valid(self):
        return all(issue.fixed or issue.severity != "ERROR" for issue in self.issues)

    def add_issue(self, category, location, message, severity="WARNING", fixed=False):
        issue = ValidationIssue(category, location, message, severity, fixed)
        self.issues.append(issue)
        if fixed:
            self.fixes.append(issue)

    def summary(self):
        total = len(self.issues)
        fixed_cnt = len(self.fixes)
        errors = sum(1 for i in self.issues if i.severity == "ERROR" and not i.fixed)
        warnings = sum(1 for i in self.issues if i.severity == "WARNING" and not i.fixed)
        return {
            "total_issues": total,
            "fixed": fixed_cnt,
            "remaining_errors": errors,
            "remaining_warnings": warnings,
            "is_valid": self.is_valid
        }


class ExcelFormatValidator:
    """Умный инспектор и авто-исправитель оформления Excel."""

    BORDER_COLOR_GRAY = "FF808080"
    DEFAULT_FONT_NAME = "Tahoma"
    DEFAULT_FONT_SIZE = 8.5
    SIG_FONT_SIZE = 11.0
    TITLE_FONT_SIZE = 11.0

    @classmethod
    def get_standard_border(cls, color=BORDER_COLOR_GRAY, style="thin"):
        side = Side(style=style, color=color)
        return Border(left=side, right=side, top=side, bottom=side)

    @classmethod
    def get_empty_border(cls):
        return Border()

    @classmethod
    def find_table_landmarks(cls, ws, total_cols=None):
        """Определяет ключевые строки таблицы: шапка (1..5), квартиры, закрытые ИПУ, Итого, Подпись."""
        landmarks = {
            "header_rows": (1, 5),
            "data_start_row": 6,
            "closed_row": None,
            "itogo_row": None,
            "sig_row": None,
            "total_cols": total_cols or ws.max_column
        }

        # Ищем строку с заголовками колонок (предыдущее/текущее/расход)
        for r in range(1, min(10, ws.max_row + 1)):
            for c in range(1, ws.max_column + 1):
                val = str(ws.cell(row=r, column=c).value or '').lower()
                if 'предыдущее' in val or 'текущее' in val or 'расход' in val:
                    landmarks["header_rows"] = (1, r)
                    landmarks["data_start_row"] = r + 1
                    break

        for r in range(landmarks["data_start_row"], ws.max_row + 1):
            val1 = str(ws.cell(row=r, column=1).value or '').strip()
            row_str = ' '.join(str(ws.cell(row=r, column=c).value or '') for c in range(1, min(10, ws.max_column + 1))).lower()
            
            if re.search(r'(?:закрытые|замен[её]нные)\s*(?:сч[её]тчики|ипу)', val1, re.IGNORECASE) or re.search(r'(?:закрытые|замен[её]нные)\s*(?:сч[её]тчики|ипу)', row_str):
                landmarks["closed_row"] = r
            elif 'итого' in val1.lower() or (val1 == '' and 'итого' in row_str):
                landmarks["itogo_row"] = r
            elif 'директор' in val1.lower() or 'директор' in row_str:
                landmarks["sig_row"] = r

        # Если total_cols не передан, найдем последний столбец с данными в шапке
        if not total_cols:
            max_c = 1
            for r in range(1, (landmarks["itogo_row"] or ws.max_row) + 1):
                for c in range(ws.max_column, 0, -1):
                    if ws.cell(row=r, column=c).value is not None:
                        if c > max_c: max_c = c
                        break
            landmarks["total_cols"] = max_c

        return landmarks

    @classmethod
    def canonical_apartment_key(cls, name):
        """
        Канонический ключ квартиры/помещения для сопоставления с шаблоном.
        """
        if not name:
            return ""
        s = str(name).strip().lower().replace('ё', 'е')
        s = re.sub(r'\s+', ' ', s)
        if re.match(r'^\d+$', s):
            return f"apt:{s}"
        m_apt = re.match(r'^(?:кв\.?|квартира)\s*([\d\w\-\/]+)$', s)
        if m_apt:
            clean_num = m_apt.group(1).replace(' ', '').replace('-', '')
            return f"apt:{clean_num}"
        m_pom = re.match(r'^(?:пом\.?|помещение)\s*([\d\w\-\/]+)$', s)
        if m_pom:
            clean_num = m_pom.group(1).replace(' ', '').replace('-', '')
            return f"pom:{clean_num}"
        m_off = re.match(r'^(?:офис)\s*([\d\w\-\/]+)$', s)
        if m_off:
            clean_num = m_off.group(1).replace(' ', '').replace('-', '')
            return f"off:{clean_num}"
        return s

    @classmethod
    def auto_fix_and_validate(cls, ws, total_cols=None, template_ws=None, sig_text=None, template_widths=None, template_apartments=None):
        """
        Гарантированное авто-исправление всех параметров оформления и валидация.
        Возвращает отчет ValidationReport. Выделяет новые строки красным шрифтом.
        """
        report = ValidationReport(filename=getattr(ws, "title", "Sheet"))
        landmarks = cls.find_table_landmarks(ws, total_cols=total_cols)
        tot_cols = landmarks["total_cols"]
        itogo_r = landmarks["itogo_row"]
        data_start_r = landmarks["data_start_row"]

        # Извлечение списка квартир из шаблона (если передан)
        tpl_apts = set()
        tpl_apt_keys = set()
        if template_apartments:
            tpl_apts = {re.sub(r'\s+', ' ', str(a).strip().lower()) for a in template_apartments if a}
            tpl_apt_keys = {cls.canonical_apartment_key(a) for a in template_apartments if a}
        elif template_ws:
            try:
                if isinstance(template_ws, str) and os.path.exists(template_ws):
                    twb = openpyxl.load_workbook(template_ws, data_only=True)
                    tws = twb.active
                elif hasattr(template_ws, 'active'):
                    tws = template_ws.active
                else:
                    tws = template_ws

                t_landmarks = cls.find_table_landmarks(tws)
                t_data_start = t_landmarks["data_start_row"]
                t_itogo = t_landmarks["itogo_row"]
                end_tr = t_itogo if t_itogo else (tws.max_row + 1)
                for tr in range(t_data_start, end_tr):
                    t_val = str(tws.cell(row=tr, column=1).value or '').strip()
                    if t_val and not any(k in t_val.lower() for k in ['итого', 'всего', 'закрытые', 'замененные', 'директор', 'подпись', 'реестр']):
                        tpl_apts.add(re.sub(r'\s+', ' ', t_val.lower()))
                        ck = cls.canonical_apartment_key(t_val)
                        if ck:
                            tpl_apt_keys.add(ck)
            except Exception:
                pass

        std_border = cls.get_standard_border()
        empty_border = cls.get_empty_border()
        empty_fill = PatternFill(fill_type=None)

        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")
        align_center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

        COLOR_RED = "FFFF0000"
        font_data = Font(name=cls.DEFAULT_FONT_NAME, size=cls.DEFAULT_FONT_SIZE, bold=False)
        font_data_strike = Font(name=cls.DEFAULT_FONT_NAME, size=cls.DEFAULT_FONT_SIZE, bold=False, strike=True)
        font_data_red = Font(name=cls.DEFAULT_FONT_NAME, size=cls.DEFAULT_FONT_SIZE, bold=False, color=COLOR_RED)
        font_data_red_strike = Font(name=cls.DEFAULT_FONT_NAME, size=cls.DEFAULT_FONT_SIZE, bold=False, strike=True, color=COLOR_RED)
        font_sig = Font(name=cls.DEFAULT_FONT_NAME, size=cls.SIG_FONT_SIZE, bold=False)

        # 1. Заголовок (строки 1 и 2)
        ws.row_dimensions[1].height = 20.0
        ws.row_dimensions[2].height = 18.0
        
        # Перепроверяем мерж строк 1 и 2
        for r_idx in (1, 2):
            for rng in list(ws.merged_cells.ranges):
                if rng.min_row == r_idx and rng.max_row == r_idx:
                    try:
                        ws.unmerge_cells(str(rng))
                    except Exception:
                        pass
            ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=tot_cols)
            cell = ws.cell(row=r_idx, column=1)
            cell.alignment = align_center
            if not cell.font or cell.font.name != "Calibri":
                cell.font = Font(name="Calibri", size=cls.TITLE_FONT_SIZE, bold=False)
            for c in range(1, tot_cols + 1):
                ws.cell(row=r_idx, column=c).border = copy(empty_border)

        # 2. Шапка таблицы (строки 3..5)
        for r in range(3, data_start_r):
            if r == 5:
                ws.row_dimensions[r].height = 27.75  # Высота для переноса двухстрочных названий
            else:
                ws.row_dimensions[r].height = 15.0

            for c in range(1, tot_cols + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = copy(std_border)
                
                # Шрифт шапки
                cur_font = cell.font
                if not cur_font or cur_font.name != cls.DEFAULT_FONT_NAME or cur_font.size not in (8.25, 8.5):
                    cell.font = copy(font_data)
                    report.add_issue("font", f"R{r}C{c}", "Исправлен шрифт шапки на Tahoma 8.5pt", fixed=True)

                # Выравнивание шапки
                val_s = str(cell.value or '')
                if '\n' in val_s or len(val_s) > 12:
                    cell.alignment = align_center_wrap
                else:
                    cell.alignment = align_center

        # 3. Тело таблицы (Строка 6 до Итого)
        end_data_r = itogo_r if itogo_r else ws.max_row
        for r in range(data_start_r, end_data_r + 1):
            val1 = str(ws.cell(row=r, column=1).value or '').strip()
            norm_val1 = re.sub(r'\s+', ' ', val1.lower())
            val1_canon = cls.canonical_apartment_key(val1)
            is_itogo_line = (r == itogo_r) or (val1.lower() == 'итого')
            is_closed_header = 'закрытые' in val1.lower()

            ws.row_dimensions[r].height = 15.0

            # Определение новой строки (появившейся после парсера/калькулятора/Аркуса, которой не было в шаблоне)
            is_new_row = False
            if not is_itogo_line and not is_closed_header and val1:
                if tpl_apt_keys:
                    if val1_canon not in tpl_apt_keys:
                        is_new_row = True
                elif tpl_apts:
                    if norm_val1 not in tpl_apts:
                        is_new_row = True
                if not is_new_row:
                    c1_cell = ws.cell(row=r, column=1)
                    if c1_cell.font and c1_cell.font.color:
                        c_rgb = str(getattr(c1_cell.font.color, 'rgb', '')).upper()
                        if c_rgb in ("FFFF0000", "FF0000", "00FF0000", "RED"):
                            is_new_row = True

            if is_new_row:
                report.add_issue(
                    "new_row",
                    f"R{r} ({val1})",
                    f"Новая строка (отсутствовала в шаблоне): «{val1}» — выделена красным шрифтом",
                    severity="INFO",
                    fixed=True
                )

            # Синхронизация зачеркивания и красного цвета по 3-колоночным блокам счетчиков (предыдущее, текущее, расход)
            if not is_itogo_line and not is_closed_header:
                for c_start in range(2, tot_cols + 1, 3):
                    c_end = min(c_start + 2, tot_cols)
                    meter_cells = [ws.cell(row=r, column=c) for c in range(c_start, c_end + 1)]
                    meter_has_red = is_new_row or any(
                        mc.font and mc.font.color and str(getattr(mc.font.color, 'rgb', '')).upper() in ("FFFF0000", "FF0000", "00FF0000", "RED")
                        for mc in meter_cells
                    )
                    meter_has_strike = any(
                        mc.font and mc.font.strike
                        for mc in meter_cells if mc.value is not None and mc.value != ""
                    )
                    for mc in meter_cells:
                        mc.border = copy(std_border)
                        is_stk = bool(mc.font and mc.font.strike) or meter_has_strike
                        if meter_has_red:
                            mc.font = copy(font_data_red_strike if is_stk else font_data_red)
                        elif is_stk:
                            mc.font = copy(font_data_strike)
                        else:
                            mc.font = copy(font_data)

                        mc.alignment = align_right
                        if mc.value is not None and mc.value != "":
                            if not str(mc.value).startswith('='):
                                mc.number_format = "##########0.#####"

            # Оформление 1-й колонки (имя квартиры / абонента)
            cell1 = ws.cell(row=r, column=1)
            cell1.border = copy(std_border)
            cell1.alignment = align_left
            cell1.number_format = "@"
            if is_new_row:
                cell1.font = copy(font_data_red)
            elif not is_itogo_line and not is_closed_header:
                cell1.font = copy(font_data)

        # 4. Промежуточные пустые строки и подпись
        if itogo_r:
            # Безопасно снимаем ВСЕ объединенные диапазоны ниже строки Итого
            for rng in list(ws.merged_cells.ranges):
                if rng.min_row > itogo_r or rng.max_row > itogo_r:
                    try:
                        ws.unmerge_cells(str(rng))
                    except Exception:
                        pass
                    if rng in ws.merged_cells.ranges:
                        try:
                            ws.merged_cells.ranges.remove(rng)
                        except Exception:
                            pass

            actual_sig_r = itogo_r + 3
            landmarks["sig_row"] = actual_sig_r

            # Очищаем промежуточные пустые строки от рамок, заливок и мусора
            for empty_r in range(itogo_r + 1, actual_sig_r):
                ws.row_dimensions[empty_r].height = 15.0
                for c in range(1, tot_cols + 1):
                    cell = ws.cell(row=empty_r, column=c)
                    if not isinstance(cell, MergedCell):
                        cell.value = None
                    cell.border = copy(empty_border)
                    cell.fill = empty_fill
                    cell.font = copy(font_data)

            # Строка подписи
            ws.row_dimensions[actual_sig_r].height = 18.0

            # Очищаем ячейки строки подписи
            for c in range(1, tot_cols + 1):
                cell = ws.cell(row=actual_sig_r, column=c)
                cell.border = copy(empty_border)
                cell.fill = empty_fill
                if c > 1 and not isinstance(cell, MergedCell):
                    cell.value = None

            # Заполняем объединенную ячейку подписи
            sig_cell = ws.cell(row=actual_sig_r, column=1)
            if sig_text and str(sig_text).strip():
                sig_cell.value = str(sig_text).strip()
            elif not sig_cell.value or 'директор' not in str(sig_cell.value).lower():
                sig_cell.value = 'Директор ООО "Южный город"  Бочарова В.М.               ____________________'
            
            sig_cell.font = copy(font_sig)
            sig_cell.alignment = Alignment(horizontal="right", vertical="center")
            ws.merge_cells(start_row=actual_sig_r, start_column=1, end_row=actual_sig_r, end_column=tot_cols)

            # 5. Физическая обрезка фантомных строк ниже подписи
            if ws.max_row > actual_sig_r:
                ws.delete_rows(actual_sig_r + 1, ws.max_row - actual_sig_r)
                report.add_issue("phantom", f"Rows > {actual_sig_r}", "Удалены фантомные строки после подписи", fixed=True)

        # 6. Физическая обрезка фантомных колонок справа
        if ws.max_column > tot_cols:
            ws.delete_cols(tot_cols + 1, ws.max_column - tot_cols)
            report.add_issue("phantom", f"Cols > {tot_cols}", "Удалены фантомные колонки справа", fixed=True)

        # 7. Ширины колонок
        if template_widths:
            for c in range(1, tot_cols + 1):
                col_letter = get_column_letter(c)
                if c in template_widths:
                    ws.column_dimensions[col_letter].width = template_widths[c]

        # 8. Настройки страницы и печати (Print Layout)
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        if tot_cols > 8:
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        else:
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

        try:
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
        except Exception:
            pass

        ws.print_title_rows = "$3:$5"
        ws.print_options.gridLines = False
        
        # Устанавливаем аккуратные поля страницы
        ws.page_margins.left = 0.4
        ws.page_margins.right = 0.4
        ws.page_margins.top = 0.6
        ws.page_margins.bottom = 0.6
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3

        new_rows_list = [issue.location for issue in report.issues if issue.category == "new_row"]
        report.stats = {
            "total_rows": ws.max_row,
            "total_cols": tot_cols,
            "itogo_row": itogo_r,
            "sig_row": landmarks.get("sig_row"),
            "orientation": ws.page_setup.orientation,
            "print_titles": ws.print_title_rows,
            "new_rows_count": len(new_rows_list),
            "new_rows": new_rows_list
        }

        return report

    @classmethod
    def validate_file_from_ws(cls, ws, filename="Workbook", template_ws=None, template_apartments=None):
        """Автономная валидация переданного worksheet с возвратом детального отчета."""
        report = ValidationReport(filename=filename)
        landmarks = cls.find_table_landmarks(ws)
        tot_cols = landmarks["total_cols"]
        itogo_r = landmarks["itogo_row"]
        sig_r = landmarks["sig_row"]
        data_start_r = landmarks["data_start_row"]

        tpl_apts = set()
        tpl_apt_keys = set()
        if template_apartments:
            tpl_apts = {re.sub(r'\s+', ' ', str(a).strip().lower()) for a in template_apartments if a}
            tpl_apt_keys = {cls.canonical_apartment_key(a) for a in template_apartments if a}
        elif template_ws:
            try:
                if isinstance(template_ws, str) and os.path.exists(template_ws):
                    twb = openpyxl.load_workbook(template_ws, data_only=True)
                    tws = twb.active
                elif hasattr(template_ws, 'active'):
                    tws = template_ws.active
                else:
                    tws = template_ws

                t_landmarks = cls.find_table_landmarks(tws)
                t_data_start = t_landmarks["data_start_row"]
                t_itogo = t_landmarks["itogo_row"]
                end_tr = t_itogo if t_itogo else (tws.max_row + 1)
                for tr in range(t_data_start, end_tr):
                    t_val = str(tws.cell(row=tr, column=1).value or '').strip()
                    if t_val and not any(k in t_val.lower() for k in ['итого', 'всего', 'закрытые', 'замененные', 'директор', 'подпись', 'реестр']):
                        tpl_apts.add(re.sub(r'\s+', ' ', t_val.lower()))
                        ck = cls.canonical_apartment_key(t_val)
                        if ck:
                            tpl_apt_keys.add(ck)
            except Exception:
                pass

        # Проверка строк таблицы, шрифтов, новых строк и рамок
        end_data_r = itogo_r if itogo_r else ws.max_row
        for r in range(data_start_r, end_data_r + 1):
            val1 = str(ws.cell(row=r, column=1).value or '').strip()
            norm_val1 = re.sub(r'\s+', ' ', val1.lower())
            val1_canon = cls.canonical_apartment_key(val1)

            is_new = False
            if val1 and not any(k in norm_val1 for k in ['итого', 'всего', 'закрытые', 'замененные']):
                if tpl_apt_keys:
                    if val1_canon not in tpl_apt_keys:
                        is_new = True
                elif tpl_apts:
                    if norm_val1 not in tpl_apts:
                        is_new = True

            if not is_new and val1 and not any(k in norm_val1 for k in ['итого', 'всего', 'закрытые', 'замененные']):
                c1_cell = ws.cell(row=r, column=1)
                if c1_cell.font and c1_cell.font.color:
                    c_rgb = str(getattr(c1_cell.font.color, 'rgb', '')).upper()
                    if c_rgb in ("FFFF0000", "FF0000", "00FF0000", "RED"):
                        is_new = True

            if is_new:
                report.add_issue("new_row", f"R{r} ({val1})", f"Строка «{val1}» отсутствовала в шаблоне (новая из Аркуса/калькулятора)", severity="INFO")
            elif val1 and not any(k in norm_val1 for k in ['итого', 'всего', 'закрытые', 'замененные']):
                # Проверка отдельных новых счетчиков, выделенных красным шрифтом
                for c_chk in range(2, tot_cols + 1, 3):
                    chk_cells = [ws.cell(row=r, column=c_chk + offset) for offset in range(3) if c_chk + offset <= tot_cols]
                    if any(mc.font and mc.font.color and str(getattr(mc.font.color, 'rgb', '')).upper() in ("FFFF0000", "FF0000", "00FF0000", "RED") for mc in chk_cells):
                        col_letter = get_column_letter(c_chk)
                        report.add_issue("new_meter", f"R{r}C{c_chk} ({val1})", f"Новые показания счетчика (колонка {col_letter}) для «{val1}» — выделены красным шрифтом", severity="INFO")

            for c in range(1, tot_cols + 1):
                cell = ws.cell(row=r, column=c)
                if cell.font and cell.font.name not in (cls.DEFAULT_FONT_NAME, "Arial", "Calibri"):
                    report.add_issue("font", f"R{r}C{c}", f"Нестандартный шрифт {cell.font.name} вместо {cls.DEFAULT_FONT_NAME}", severity="WARNING")
                
                if not cell.border or not cell.border.top or not cell.border.top.style:
                    report.add_issue("border", f"R{r}C{c}", "Отсутствует рамка у ячейки таблицы", severity="WARNING")
                elif cell.border.top.color and str(getattr(cell.border.top.color, 'rgb', '')) not in (cls.BORDER_COLOR_GRAY, "00000000", ""):
                    report.add_issue("border", f"R{r}C{c}", f"Цвет рамки отличается от серого #808080", severity="INFO")

        # Проверка пустых строк перед подписью
        if itogo_r and sig_r:
            for empty_r in range(itogo_r + 1, sig_r):
                for c in range(1, tot_cols + 1):
                    cell = ws.cell(row=empty_r, column=c)
                    if cell.border and cell.border.top and cell.border.top.style:
                        report.add_issue("border", f"R{empty_r}C{c}", "Лишняя рамка в пустой строке перед подписью", severity="ERROR")

        # Проверка параметров печати
        if ws.print_title_rows not in ("$3:$5", "3:5", "$1:$5", "1:5"):
            report.add_issue("print", "PageSetup", f"Не установлены сквозные строки печати ($3:$5), текущее: {ws.print_title_rows}", severity="WARNING")

        return report

    @classmethod
    def validate_file(cls, filepath, template_path=None):
        """Автономная валидация любого xlsx файла с возвратом детального отчета."""
        wb = openpyxl.load_workbook(filepath, data_only=False)
        ws = wb.active
        return cls.validate_file_from_ws(ws, filename=os.path.basename(filepath), template_ws=template_path)

