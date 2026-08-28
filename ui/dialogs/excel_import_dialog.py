import os
import json
from typing import Dict, List, Any

from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit,
    QPushButton, QTableWidget, QTableWidgetItem, QHeaderView,
    QComboBox, QFileDialog, QSizePolicy, QMenu, QAbstractItemView,
    QWidget, QFrame
)
from PySide6.QtCore import Qt, Signal, QSettings, QPoint, QSize
from PySide6.QtGui import QColor, QBrush, QFont, QPainter, QPainterPath, QPen, QLinearGradient
from openpyxl.utils import get_column_letter

# Assuming ExcelManager exists in core.excel_parser
try:
    from core.excel_parser import ExcelManager
except ImportError:
    class ExcelManager:
        @staticmethod
        def preview_worksheet(path, sheet_name=None):
            return ["Sheet1"], [["Адрес", "ХВС", "ГВС", "ДОБ"], ["Душистая 45", 10.0, 5.0, 1.0]]
        @staticmethod
        def detect_column_roles(rows_data):
            return {0: 'house', 1: 'hvs', 2: 'gvs', 3: 'dob'}
        @staticmethod
        def extract_values_by_mapping(rows_data, mapping, house_name):
            return {"hvs": 312.5, "gvs": 104.2, "dob": 18.3}

from ui.styles import ThemeManager, get_svg_icon


class ExcelImportDialog(QDialog):
    values_accepted = Signal(float, float, float)

    def __init__(self, parent=None, house_name="", initial_path=""):
        super().__init__(parent)
        self.house_name = house_name
        self.initial_path = initial_path
        
        self.setWindowTitle("📥 Импорт показаний из таблицы")
        self.setWindowFlags(
            Qt.WindowType.FramelessWindowHint
            | Qt.WindowType.Dialog
            | Qt.WindowType.WindowStaysOnTopHint
        )
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground, True)
        self.resize(850, 550)
        self.setModal(True)

        self.mapping = {}  # col_idx -> role
        self.rows_data = []
        self.sheets = []
        self._drag_pos = None
        self.current_found_values = None
        
        self.init_ui()
        
        if self.initial_path and os.path.exists(self.initial_path):
            self.txt_file.setText(self.initial_path)
            self._load_file(self.initial_path)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        rect = self.rect().adjusted(1, 1, -1, -1)
        theme_name = ThemeManager.get_current_theme_name()
        accent = ThemeManager.get_current_accent_color(theme_name)
        is_light = theme_name in ("Pearl Light", "Как дома")

        path = QPainterPath()
        radius = 2 if theme_name == "Как дома" else 14
        path.addRoundedRect(rect, radius, radius)

        if theme_name == "Как дома":
            painter.setBrush(QBrush(QColor("#ECE9D8")))
            painter.setPen(QPen(QColor("#7F9DB9"), 1))
        elif is_light:
            grad = QLinearGradient(0, 0, 0, rect.height())
            grad.setColorAt(0.0, QColor(255, 255, 255, 255))
            grad.setColorAt(1.0, QColor(248, 250, 252, 252))
            border_color = QColor(accent)
            border_color.setAlpha(200)
            glow_pen = QPen(border_color, 1.5)
            painter.setPen(glow_pen)
            painter.setBrush(QBrush(grad))
        else:
            grad = QLinearGradient(0, 0, 0, rect.height())
            grad.setColorAt(0.0, QColor(18, 28, 50, 240))
            grad.setColorAt(1.0, QColor(15, 23, 42, 240))
            border_color = QColor(accent)
            border_color.setAlpha(180)
            glow_pen = QPen(border_color, 1.5)
            painter.setPen(glow_pen)
            painter.setBrush(QBrush(grad))

        painter.drawPath(path)

    def init_ui(self):
        root_layout = QVBoxLayout(self)
        root_layout.setContentsMargins(20, 16, 20, 18)
        root_layout.setSpacing(12)

        theme_name = ThemeManager.get_current_theme_name()
        accent = ThemeManager.get_current_accent_color(theme_name)
        is_light = theme_name in ("Pearl Light", "Как дома")

        # Title bar
        header_row = QHBoxLayout()
        header_row.setSpacing(10)
        
        self.lbl_main_title = QLabel("📥 Импорт показаний из таблицы")
        title_color = "#0A246A" if theme_name == "Как дома" else ("#0F172A" if is_light else "#F8FAFC")
        self.lbl_main_title.setStyleSheet(f"font-size: 18px; font-weight: bold; color: {title_color}; background: transparent;")

        header_row.addWidget(self.lbl_main_title, 1)

        self.btn_close = QPushButton("✕")
        self.btn_close.setFixedSize(30, 30)
        self.btn_close.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_close.setStyleSheet("""
            QPushButton {
                background: rgba(255, 255, 255, 0.06);
                color: #94A3B8;
                border: 1px solid rgba(255, 255, 255, 0.12);
                border-radius: 15px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: rgba(239, 68, 68, 0.25);
                color: #EF4444;
                border: 1px solid rgba(239, 68, 68, 0.5);
            }
        """)
        self.btn_close.clicked.connect(self.reject)
        header_row.addWidget(self.btn_close)

        root_layout.addLayout(header_row)

        # 1. File Selection Row
        file_row = QHBoxLayout()
        lbl_file = QLabel("Файл:")
        lbl_file.setStyleSheet(f"color: {title_color}; font-weight: bold; font-size: 13px; background: transparent;")
        
        self.txt_file = QLineEdit()
        self.txt_file.setReadOnly(True)
        self.txt_file.setMinimumHeight(34)
        
        self.btn_browse = QPushButton("Обзор")
        self.btn_browse.setMinimumHeight(34)
        self.btn_browse.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_browse.clicked.connect(self._on_browse)
        
        self.btn_folder = QPushButton("📂")
        self.btn_folder.setMinimumHeight(34)
        self.btn_folder.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_folder.clicked.connect(self._on_browse)
        
        file_row.addWidget(lbl_file)
        file_row.addWidget(self.txt_file, 1)
        file_row.addWidget(self.btn_browse)
        file_row.addWidget(self.btn_folder)
        root_layout.addLayout(file_row)

        # 2. Sheet Selector Row
        self.sheet_row = QWidget()
        sheet_layout = QHBoxLayout(self.sheet_row)
        sheet_layout.setContentsMargins(0, 0, 0, 0)
        lbl_sheet = QLabel("Лист:")
        lbl_sheet.setStyleSheet(f"color: {title_color}; font-weight: bold; font-size: 13px; background: transparent;")
        
        self.combo_sheet = QComboBox()
        self.combo_sheet.setMinimumHeight(34)
        self.combo_sheet.currentIndexChanged.connect(self._on_sheet_changed)
        
        sheet_layout.addWidget(lbl_sheet)
        sheet_layout.addWidget(self.combo_sheet, 1)
        self.sheet_row.setVisible(False)
        root_layout.addWidget(self.sheet_row)

        # 3. Info Label
        self.lbl_info = QLabel(f"🏢 Текущий дом: {self.house_name}\n✅ Столбцы не назначены")
        self.lbl_info.setStyleSheet(f"color: {title_color}; font-size: 13px; background: transparent;")
        root_layout.addWidget(self.lbl_info)

        # 4. Table Preview
        self.table = QTableWidget()
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.horizontalHeader().setSectionsClickable(True)
        self.table.horizontalHeader().sectionClicked.connect(self._on_column_clicked)
        self.table.setAlternatingRowColors(True)
        
        if theme_name == "Как дома":
            self.table.setStyleSheet("""
                QTableWidget { background: white; border: 1px solid #7F9DB9; gridline-color: #7F9DB9; color: black; }
                QHeaderView::section { background: #ECE9D8; color: black; border: 1px solid #7F9DB9; font-weight: bold; padding: 4px; }
            """)
        else:
            self.table.setStyleSheet("""
                QTableWidget {
                    background-color: transparent;
                    alternate-background-color: rgba(255, 255, 255, 0.03);
                    border: 1px solid rgba(255, 255, 255, 0.1);
                    border-radius: 8px;
                    gridline-color: rgba(255, 255, 255, 0.05);
                }
                QHeaderView::section {
                    background-color: rgba(255, 255, 255, 0.08);
                    padding: 6px;
                    border: none;
                    border-bottom: 1px solid rgba(255, 255, 255, 0.1);
                    border-right: 1px solid rgba(255, 255, 255, 0.05);
                    font-weight: bold;
                }
            """)
        
        root_layout.addWidget(self.table, 1)

        # 5. Results Preview Row
        self.lbl_results = QLabel("⚠️ Дом не найден в таблице")
        self.lbl_results.setStyleSheet(f"color: {title_color}; font-weight: bold; font-size: 14px; background: transparent;")
        self.lbl_results.setAlignment(Qt.AlignmentFlag.AlignCenter)
        root_layout.addWidget(self.lbl_results)

        # 6. Bottom Buttons Row
        btn_box = QHBoxLayout()
        btn_box.addStretch()
        
        self.btn_cancel = QPushButton("Отмена")
        self.btn_cancel.setMinimumHeight(36)
        self.btn_cancel.setMinimumWidth(100)
        self.btn_cancel.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_cancel.clicked.connect(self.reject)
        
        self.btn_apply = QPushButton("✨ Применить значения")
        self.btn_apply.setMinimumHeight(36)
        self.btn_apply.setMinimumWidth(180)
        self.btn_apply.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_apply.setEnabled(False)
        self.btn_apply.clicked.connect(self._on_apply)
        
        if theme_name == "Как дома":
            classic_btn = "background: #ECE9D8; color: #000; border: 1px solid #7F9DB9; padding: 5px;"
            self.btn_cancel.setStyleSheet(classic_btn)
            self.btn_apply.setStyleSheet("background: #0A246A; color: white; border: 1px solid #0A246A; font-weight: bold; padding: 5px;")
            self.btn_browse.setStyleSheet(classic_btn)
            self.btn_folder.setStyleSheet(classic_btn)
        else:
            self.btn_cancel.setObjectName("SecondaryButton")
            self.btn_apply.setObjectName("PrimaryButton")
            self.btn_browse.setObjectName("SecondaryButton")
            self.btn_folder.setObjectName("SecondaryButton")

        btn_box.addWidget(self.btn_cancel)
        btn_box.addWidget(self.btn_apply)
        root_layout.addLayout(btn_box)

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            pos = event.position().toPoint() if hasattr(event, 'position') else event.pos()
            child = self.childAt(pos)
            is_interactive = False
            w = child
            while w and w is not self:
                if isinstance(w, (QLineEdit, QPushButton, QComboBox, QTableWidget)):
                    is_interactive = True
                    break
                w = w.parentWidget()
            if not is_interactive:
                self._drag_pos = event.globalPosition().toPoint() if hasattr(event, 'globalPosition') else event.globalPos()
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if self._drag_pos and event.buttons() & Qt.MouseButton.LeftButton:
            curr_pos = event.globalPosition().toPoint() if hasattr(event, 'globalPosition') else event.globalPos()
            delta = curr_pos - self._drag_pos
            self.move(self.pos() + delta)
            self._drag_pos = curr_pos
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        self._drag_pos = None
        super().mouseReleaseEvent(event)

    def _on_browse(self):
        path, _ = QFileDialog.getOpenFileName(self, "Выберите Excel файл", "", "Excel (*.xlsx *.xls)")
        if path:
            self.txt_file.setText(path)
            self._load_file(path)

    def _load_file(self, path, sheet_name=None):
        try:
            self.sheets, self.rows_data, active_name = ExcelManager.preview_worksheet(path, sheet_name, max_rows=None)
        except Exception as e:
            self.lbl_results.setText(f"❌ Ошибка загрузки: {str(e)}")
            return

        if not sheet_name:
            self.combo_sheet.blockSignals(True)
            self.combo_sheet.clear()
            self.combo_sheet.addItems(self.sheets)
            if active_name in self.sheets:
                self.combo_sheet.setCurrentIndex(self.sheets.index(active_name))
            self.combo_sheet.blockSignals(False)
            self.sheet_row.setVisible(len(self.sheets) > 1)

        # Try to restore saved mapping for this file
        basename = os.path.basename(path)
        settings = QSettings("WaterMetrics", "ImportMappings")
        saved = settings.value(basename, "", type=str)
        if saved:
            try:
                loaded = json.loads(saved)
                self.mapping = {int(k): v for k, v in loaded.items()}
            except Exception:
                self.mapping = {}
        else:
            self.mapping = {}

        # Auto-detect mapping if nothing was restored
        if not self.mapping:
            try:
                detected = ExcelManager.detect_column_roles(self.rows_data)
                # detected is {role_name -> col_idx}, convert to {col_idx -> role_code}
                role_map = {'дом': 'house', 'хвс': 'hvs', 'гвс': 'gvs', 'доб': 'dob'}
                self.mapping = {col_idx: role_map.get(role, role) for role, col_idx in detected.items()}
            except Exception:
                self.mapping = {}

        self._update_table()
        self._search_house_values()

    def _on_sheet_changed(self):
        if self.txt_file.text():
            self._load_file(self.txt_file.text(), self.combo_sheet.currentText())

    def _update_table(self):
        if not self.rows_data:
            return

        num_cols = max(len(row) for row in self.rows_data)
        self.table.setColumnCount(num_cols)
        self.table.setRowCount(len(self.rows_data))

        for r_idx, row in enumerate(self.rows_data):
            for c_idx, val in enumerate(row):
                item = QTableWidgetItem(str(val) if val is not None else "")
                self.table.setItem(r_idx, c_idx, item)

        self._update_headers()

    def _update_headers(self):
        role_icons = {'house': '🏠', 'hvs': '💧', 'gvs': '🔥', 'dob': '➕'}

        for c_idx in range(self.table.columnCount()):
            letter = get_column_letter(c_idx + 1)
            role = self.mapping.get(c_idx)
            text = f"{letter} {role_icons[role]}" if role in role_icons else letter
            item = QTableWidgetItem(text)
            self.table.setHorizontalHeaderItem(c_idx, item)

        # Highlight mapped columns
        theme_name = ThemeManager.get_current_theme_name()
        accent = ThemeManager.get_current_accent_color(theme_name)
        bg_color = QColor(accent)
        bg_color.setAlpha(25)

        for c_idx in range(self.table.columnCount()):
            is_mapped = c_idx in self.mapping
            for r_idx in range(self.table.rowCount()):
                item = self.table.item(r_idx, c_idx)
                if item:
                    if is_mapped:
                        item.setBackground(QBrush(bg_color))
                    else:
                        item.setBackground(QBrush(Qt.GlobalColor.transparent))

        # Update info label
        parts = []
        for c, r in self.mapping.items():
            col_letter = get_column_letter(c + 1)
            if r == 'house': parts.append(f"Дом({col_letter})")
            elif r == 'hvs': parts.append(f"ХВС({col_letter})")
            elif r == 'gvs': parts.append(f"ГВС({col_letter})")
            elif r == 'dob': parts.append(f"ДОБ({col_letter})")

        status = "✅ Столбцы назначены: " + ", ".join(parts) if parts else "💡 Кликните по заголовкам столбцов (A, B, C...) для назначения"
        self.lbl_info.setText(f"🏢 Текущий дом: {self.house_name}\n{status}")

    def _on_column_clicked(self, col_idx):
        menu = QMenu(self)

        a_house = menu.addAction("🏠 Дом / Адрес")
        a_hvs = menu.addAction("💧 ХВС (холодная вода)")
        a_gvs = menu.addAction("🔥 ГВС (горячая вода)")
        a_dob = menu.addAction("➕ ДОБ. (добавочные)")
        menu.addSeparator()
        a_clear = menu.addAction("— Не использовать")

        action = menu.exec(self.cursor().pos())
        if action == a_house: self._assign_role(col_idx, 'house')
        elif action == a_hvs: self._assign_role(col_idx, 'hvs')
        elif action == a_gvs: self._assign_role(col_idx, 'gvs')
        elif action == a_dob: self._assign_role(col_idx, 'dob')
        elif action == a_clear: self._assign_role(col_idx, None)

    def _assign_role(self, col_idx, role):
        if role is None:
            if col_idx in self.mapping:
                del self.mapping[col_idx]
        else:
            # Clear this role from other columns
            self.mapping = {k: v for k, v in self.mapping.items() if v != role}
            self.mapping[col_idx] = role

        self._update_headers()
        self._search_house_values()

    def _search_house_values(self):
        if not self.rows_data:
            self.lbl_results.setText("⚠️ Таблица пуста")
            self.btn_apply.setEnabled(False)
            return

        if 'hvs' not in self.mapping.values() and 'gvs' not in self.mapping.values():
            self.lbl_results.setText("⚠️ Назначьте столбцы ХВС или ГВС")
            self.btn_apply.setEnabled(False)
            return

        try:
            # Convert col_idx->role to role->col_idx for extract_values_by_mapping
            role_to_col = {}
            role_remap = {'house': 'дом', 'hvs': 'хвс', 'gvs': 'гвс', 'dob': 'доб'}
            for col_idx, role in self.mapping.items():
                ru_role = role_remap.get(role, role)
                role_to_col[ru_role] = col_idx

            res = ExcelManager.extract_values_by_mapping(self.rows_data, role_to_col, self.house_name)
            if res:
                self.current_found_values = res
                hvs = res.get('хвс', 0.0)
                gvs = res.get('гвс', 0.0)
                dob = res.get('доб', 0.0)
                row_idx = res.get('row_idx')

                theme_name = ThemeManager.get_current_theme_name()
                accent = ThemeManager.get_current_accent_color(theme_name)
                is_light = theme_name in ("Pearl Light", "Как дома")
                val_color = "#0A246A" if theme_name == "Как дома" else ("#028090" if is_light else accent)

                text = f"ХВС: {hvs}  |  ГВС: {gvs}  |  ДОБ: {dob}"
                if row_idx is not None:
                    text += f"  (строка {row_idx + 1})"
                self.lbl_results.setText(text)
                self.lbl_results.setStyleSheet(f"color: {val_color}; font-weight: bold; font-size: 15px; background: transparent;")
                self.btn_apply.setEnabled(True)

                # Scroll to and highlight matched row
                if row_idx is not None and row_idx < self.table.rowCount():
                    hl_color = QColor(accent)
                    hl_color.setAlpha(70)
                    for c_idx in range(self.table.columnCount()):
                        item = self.table.item(row_idx, c_idx)
                        if item:
                            item.setBackground(QBrush(hl_color))
                    first_item = self.table.item(row_idx, 0)
                    if first_item:
                        self.table.scrollToItem(first_item, QAbstractItemView.ScrollHint.PositionAtCenter)
            else:
                self.lbl_results.setText(f"⚠️ Дом «{self.house_name}» не найден (или в строке нет числовых данных)")
                self.btn_apply.setEnabled(False)
                self.current_found_values = None

        except Exception as e:
            self.lbl_results.setText(f"❌ Ошибка поиска: {str(e)}")
            self.btn_apply.setEnabled(False)
            self.current_found_values = None

    def _on_apply(self):
        if self.current_found_values:
            # Save mapping settings and store session cache
            if self.txt_file.text():
                basename = os.path.basename(self.txt_file.text())
                settings = QSettings("WaterMetrics", "ImportMappings")
                settings.setValue(basename, json.dumps({str(k): v for k, v in self.mapping.items()}))

                # Сохраняем в кэш сессии для авто-подстановки при переключении домов
                try:
                    from services.import_session_service import ImportSessionService
                    role_to_col = {}
                    role_remap = {'house': 'дом', 'hvs': 'хвс', 'gvs': 'гвс', 'dob': 'доб'}
                    for col_idx, role in self.mapping.items():
                        ru_role = role_remap.get(role, role)
                        role_to_col[ru_role] = col_idx

                    sheet_title = self.combo_sheet.currentText() if hasattr(self, 'combo_sheet') and self.combo_sheet.count() > 0 else ""
                    ImportSessionService.set_session_import(
                        file_path=self.txt_file.text(),
                        sheet_name=sheet_title,
                        rows_data=self.rows_data,
                        mapping=role_to_col
                    )
                except Exception:
                    pass

            hvs = float(self.current_found_values.get('хвс') or 0.0)
            gvs = float(self.current_found_values.get('гвс') or 0.0)
            dob = float(self.current_found_values.get('доб') or 0.0)
            self.values_accepted.emit(hvs, gvs, dob)
            self.accept()
