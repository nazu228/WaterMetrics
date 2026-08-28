"""
services/folder_service.py — Интеллектуальный навигатор по структуре папок, месяцам и домам (WaterMetrics).
Обеспечивает:
1. Автоматический поиск папки следующего месяца (напр. '08 Август 2026' при шаблоне из '07 Июль 2026').
2. Умный поиск файла Аркус в целевом месяце с поддержкой аббревиатур (душ 45, 2-я цел 11, пос 28 и т.д.).
3. Проверку наличия уже сформированного отчета в целевой папке (включая подпапку ГОТОВО).
4. Проверку блокировки итогового файла в Microsoft Excel (PermissionError).
5. Кросс-валидацию совпадения дома и лицевых счетов между Шаблоном и Аркусом.
"""

import os
import re
import time
from datetime import datetime
from typing import Dict, List, Optional, Tuple

MONTHS_RU = [
    "январь", "февраль", "март", "апрель", "май", "июнь",
    "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь"
]

MONTH_NAMES_CAP = [
    "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
]


class FolderNavigationService:
    """Сервис умной маршрутизации и защиты файлов."""

    @classmethod
    def clean_ordinals(cls, text: str) -> str:
        """Нормализует порядковые числительные (2-я, 2я, 3-я -> 2, 3)."""
        return re.sub(r'(\d+)[\-\_]?(?:я|ая|й|ий|ый|е|го)\b', r'\1', text, flags=re.IGNORECASE)

    @classmethod
    def normalize_house_key(cls, name: str) -> Tuple[List[str], List[str]]:
        """Извлекает ключевые корни слов и номера домов/буквы для надежного сопоставления."""
        s = str(name or '').lower().strip()
        s = s.replace('ё', 'е').replace('.xlsx', '').replace('.xls', '')

        # 1. Транслитерация латинских букв в индексах домов в кириллицу (например '6a'/'6A' -> '6а')
        lat_to_cyr = {'a': 'а', 'b': 'б', 'c': 'с', 'k': 'к', 'm': 'м', 'h': 'н', 'p': 'р', 'x': 'х', 'e': 'е', 'o': 'о', 'y': 'у'}
        s = re.sub(r'(\d+)\s*([a-z])\b', lambda m: m.group(1) + lat_to_cyr.get(m.group(2), m.group(2)), s)

        # 2. Склеивание буквенных индексов с номерами: '6 - а' -> '6а', '6/а' -> '6а', '6 а' -> '6а'
        s = re.sub(r'(\d+)\s*[\-\/\\]\s*([а-яa-z])\b', r'\1\2', s)
        s = re.sub(r'(\d+)\s+([а-яa-z])\b', r'\1\2', s)

        # 3. Разделение слитно написанных улиц и номеров: 'посадского32' -> 'посадского 32', 'гассия6а' -> 'гассия 6а'
        s = re.sub(r'([а-яa-z])(\d+)', r'\1 \2', s)

        # 4. Удаление 4-значных годов, дат (05.26, 08.2026) и месяцев с годами (май 26, август 2026)
        s = re.sub(r'\b20\d{2}\b', ' ', s)
        s = re.sub(r'\b(0[1-9]|1[0-2])[\.\-_](?:20\d{2}|\d{2})\b', ' ', s)
        s = re.sub(r'\b(янв\w*|фев\w*|мар\w*|апр\w*|май\w*|июн\w*|июл\w*|авг\w*|сен\w*|окт\w*|ноя\w*|дек\w*)(?:\s*(?:20\d{2}|\d{2}))?\b', ' ', s)

        # 5. Очистка стандартных префиксов и знаков пунктуации
        s = re.sub(r'^(?:г\.\s*краснодар,?\s*)?(?:ул\.(?:\s*им\.(?:\s*героя|\s*валерия)?(?:\s*а\.а\.)?|\s*им\.)?\s*)?', ' ', s)
        s = re.sub(r'^(?:по\s+дому|по\s+адресу|объект|дом|ул\.)\s*:?\s*', ' ', s)
        s = re.sub(r'[\+\-\_\,\.\(\)\;\:\/]', ' ', s)

        # 6. Удаление служебных стоп-слов
        s = re.sub(r'\b(тест|test|копия|шаблон|сверка|аркус|арк|arcus|arc|выгрузка|выгрузки|выгруз|юг|юд|отчет|мкд|дом|г|гсделанный|ул|улица)\b', ' ', s)

        # 7. Стандартизация популярных сокращений улиц
        replacements = [
            (r'\bпосадского\b', 'пос'),
            (r'\bпосадск\w*\b', 'пос'),
            (r'\bпосад\b', 'пос'),
            (r'\bцелиноградская\b', 'цел'),
            (r'\bцелиноград\w*\b', 'цел'),
            (r'\bдубравная\b', 'дуб'),
            (r'\bдубравн\w*\b', 'дуб'),
            (r'\bдушистая\b', 'душ'),
            (r'\bдушист\w*\b', 'душ'),
            (r'\bзеленоградская\b', 'зел'),
            (r'\bзеленоград\w*\b', 'зел'),
            (r'\bаверкиева\b', 'авер'),
            (r'\bаверкиев\w*\b', 'авер'),
            (r'\bаверки\w*\b', 'авер'),
            (r'\bтрошева\b', 'трошев'),
            (r'\bтрошев\w*\b', 'трошев'),
            (r'\bгассия\b', 'гасс'),
            (r'\bгассий\w*\b', 'гасс'),
            (r'\bгасси\w*\b', 'гасс'),
            (r'\bкореновская\b', 'корен'),
            (r'\bкоренов\w*\b', 'корен'),
            (r'\bтепличная\b', 'тепл'),
            (r'\bтепличн\w*\b', 'тепл'),
            (r'\bкомандорская\b', 'команд'),
            (r'\bкомандор\w*\b', 'команд'),
            (r'\bягодина\b', 'ягод'),
            (r'\bягодин\w*\b', 'ягод'),
            (r'\bсоколова\b', 'сокол'),
            (r'\bсоколов\w*\b', 'сокол'),
            (r'\bкрасных\s+партизан\b', 'партизан'),
            (r'\bпартизан\w*\b', 'партизан'),
            (r'\bтургенева\b', 'турген'),
            (r'\bтургенев\w*\b', 'турген'),
            (r'\bлузана\b', 'лузан'),
        ]
        for pattern, repl in replacements:
            s = re.sub(pattern, repl, s)

        # Нормализация порядковых приставок к улицам (например '2-я Целиноградская' / '1-я Целиноградская' -> 'цел')
        s = re.sub(r'\b[1-9]\s*(?:я|ая|е|ье|ий|ой|й)?\s*(цел|пос|душ|дуб|зел|авер|трошев|гасс|корен|тепл|команд|ягод|сокол)\b', r'\1', s)

        # Извлекаем числа с индексами (например: 1, 11, 7, 6а, 34, 45, 28)
        digits = re.findall(r'\d+[а-я]?', s)

        # Извлекаем значимые корни слов (длиной >= 3, исключая стоп-слова)
        words = [w for w in re.findall(r'[а-я]+', s) if len(w) >= 3 and w not in ('дом', 'кв', 'лит', 'корп', 'стр', 'ул')]

        return words, digits

    @classmethod
    def is_house_match(cls, name1: str, name2: str) -> bool:
        """Проверяет соответствие двух наименований домов (например 'Гассия 6а' и 'ул. им. Валерия Гассия, 6А')."""
        w1, d1 = cls.normalize_house_key(name1)
        w2, d2 = cls.normalize_house_key(name2)

        # Номера домов должны быть определены
        if not d1 or not d2:
            return False

        # 1. Проверка совпадения номеров домов (с учетом буквенных индексов и без них)
        digits_matched = False
        if set(d1) == set(d2):
            digits_matched = True
        else:
            # Сравнение базовых чисел (например '6' и '6а')
            base1 = [re.sub(r'[а-я]', '', x) for x in d1]
            base2 = [re.sub(r'[а-я]', '', x) for x in d2]
            if set(base1) == set(base2):
                l1 = [re.findall(r'[а-я]', x) for x in d1]
                l2 = [re.findall(r'[а-я]', x) for x in d2]
                flat_l1 = [l for sub in l1 for l in sub]
                flat_l2 = [l for sub in l2 for l in sub]
                # Если в обоих указаны разные буквы ('6а' и '6б') — не совпадение
                if not (flat_l1 and flat_l2 and set(flat_l1) != set(flat_l2)):
                    digits_matched = True

        if not digits_matched:
            return False

        # Если одно наименование содержит слова улицы, а другое — только цифры (напр. номер строки '1'), это не совпадение
        if bool(w1) != bool(w2):
            return False

        # Если оба без слов улицы (чисто номера домов)
        if not w1 and not w2:
            return True

        # Хотя бы одно слово/корень должно совпадать
        for word1 in w1:
            for word2 in w2:
                if word1 in word2 or word2 in word1:
                    return True
        return False

    @classmethod
    def extract_month_and_year_from_path(cls, path: str) -> Tuple[Optional[int], Optional[int]]:
        """Определяет номер месяца (1..12) и год из пути к файлу или папке."""
        if not path:
            return None, None

        norm = path.replace('\\', '/')
        parts = norm.split('/')

        # 1. Проверяем папки снизу вверх
        for part in reversed(parts):
            # Проверка формата "07 Июль 2026" или "07_Июль_2026" или "Июль 2026"
            for m_idx, m_name in enumerate(MONTHS_RU, start=1):
                if m_name in part.lower():
                    m_year = re.search(r'\b(20\d{2})\b', part)
                    year = int(m_year.group(1)) if m_year else datetime.now().year
                    return m_idx, year

            # Проверка формата "07.2026" или "07-2026"
            m_dig = re.search(r'\b(0[1-9]|1[0-2])[\.\-_](20\d{2})\b', part)
            if m_dig:
                return int(m_dig.group(1)), int(m_dig.group(2))

        return None, None

    _arcus_addr_cache: Dict[str, Tuple[float, str]] = {}

    @classmethod
    def extract_house_name_from_arcus_content(cls, path: str) -> str:
        """Извлекает адрес дома из внутренних ячеек выгрузки Аркус (строки 1..5) с мгновенным кэшированием."""
        if not path or not os.path.isfile(path):
            return ""
        try:
            mtime = os.path.getmtime(path)
            if path in cls._arcus_addr_cache:
                cached_mtime, cached_addr = cls._arcus_addr_cache[path]
                if cached_mtime == mtime:
                    return cached_addr

            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            res = ""
            for row in ws.iter_rows(min_row=1, max_row=6, max_col=10, values_only=True):
                for val in row:
                    if not val:
                        continue
                    s = str(val).strip()
                    if any(kw in s.lower() for kw in ['краснодар', 'ул.', 'д.', 'улица', 'дом']):
                        clean = re.sub(r'^(?:г\.\s*Краснодар,?\s*)?(?:ул\.(?:\s*им\.\s*Героя)?\s*)?', '', s, flags=re.IGNORECASE)
                        clean = re.sub(r',?\s*д\.\s*', ' ', clean).strip()
                        res = clean
                        break
                if res:
                    break
            wb.close()
            cls._arcus_addr_cache[path] = (mtime, res)
            return res
        except Exception:
            return ""

    @classmethod
    def find_smart_arcus_path(
        cls,
        tpl_filename: str,
        current_month_dir: str,
        next_m: int,
        next_y: int,
        curr_m: int,
        template_path: str = "",
        house_name: str = ""
    ) -> Tuple[Optional[str], Optional[str]]:
        """
        Интеллектуальный многоуровневый поиск файла Аркус:
        - Проверяет структуру на 2-3 папки вверх (current_month -> parent1 -> parent2 -> parent3).
        - Ищет целевую папку следующего месяца.
        - Сканирует подпапки Аркус / Выгрузки / Исходные данные с глубиной до 2 уровней.
        - Быстро сопоставляет по имени дома и кэшированному содержимому.
        - Ранжирует кандидатов по качеству соответствия дому и свежести.
        Возвращает: (найденный_путь_аркус, папка_следующего_месяца)
        """
        if not current_month_dir or not os.path.exists(current_month_dir):
            return None, None

        if not house_name and template_path and os.path.exists(template_path):
            from core.excel_parser import ExcelManager
            house_name = ExcelManager.extract_house_name(template_path)

        match_targets = [t for t in [house_name, tpl_filename] if t]
        if not match_targets:
            return None, None

        next_m_str = f"{next_m:02d}"
        next_m_name = MONTHS_RU[next_m - 1]
        curr_m_str = f"{curr_m:02d}"
        curr_m_name = MONTHS_RU[curr_m - 1]

        # Уровни иерархии (до 3 папок вверх: 1 папка = root, 2 папки = grandparent, 3 папки = great_grandparent)
        parent1 = os.path.dirname(current_month_dir)
        parent2 = os.path.dirname(parent1) if parent1 else None
        parent3 = os.path.dirname(parent2) if parent2 else None

        parents = [p for p in [parent1, parent2, parent3] if p and os.path.isdir(p)]

        # 1. Поиск папки следующего месяца
        next_month_dir = None
        for p in parents:
            for entry in os.listdir(p):
                full_p = os.path.join(p, entry)
                if os.path.isdir(full_p):
                    e_low = entry.lower()
                    if (next_m_str in e_low and next_m_name in e_low) or (next_m_name in e_low and str(next_y) in e_low):
                        next_month_dir = full_p
                        break
                    elif next_m_name in e_low and not next_month_dir:
                        next_month_dir = full_p
            if next_month_dir:
                break

        # 2. Сбор потенциальных директорий для поиска Аркуса
        search_dirs = []

        # Целевой месяц и его подпапки
        if next_month_dir and os.path.exists(next_month_dir):
            search_dirs.append(next_month_dir)
            try:
                for root, dirs, _ in os.walk(next_month_dir):
                    rel_depth = len(os.path.relpath(root, next_month_dir).split(os.sep))
                    if rel_depth <= 2 and root not in search_dirs:
                        search_dirs.append(root)
            except Exception:
                pass

        # Поиск папок 'Аркус' / 'Выгрузка' на уровнях parent1 и parent2 (до 2 папок вверх)
        keywords = ('аркус', 'arcus', 'выгрузк', 'выгрузка', 'показан', 'ипу', 'ipu', next_m_name)
        for p in parents:
            try:
                for entry in os.listdir(p):
                    full_p = os.path.join(p, entry)
                    if os.path.isdir(full_p):
                        e_low = entry.lower()
                        if any(kw in e_low for kw in keywords):
                            if full_p not in search_dirs:
                                search_dirs.append(full_p)
                            try:
                                for sub in os.listdir(full_p):
                                    sub_p = os.path.join(full_p, sub)
                                    if os.path.isdir(sub_p) and sub_p not in search_dirs:
                                        search_dirs.append(sub_p)
                            except Exception:
                                pass
            except Exception:
                pass

        # Текущий месяц и его подпапки
        if current_month_dir not in search_dirs:
            search_dirs.append(current_month_dir)
            for sub in ("Аркус", "аркус", "Arcus", "arcus", "Выгрузка", "выгрузка"):
                sub_p = os.path.join(current_month_dir, sub)
                if os.path.isdir(sub_p) and sub_p not in search_dirs:
                    search_dirs.append(sub_p)

        # 3. Поиск и ранжирование подходящих файлов Аркус (сначала быстрый поиск по имени)
        candidates = []
        seen_paths = set()

        for s_dir in search_dirs:
            if not os.path.isdir(s_dir):
                continue
            try:
                for f in os.listdir(s_dir):
                    if f.endswith(('.xlsx', '.xls')) and not f.startswith('~$'):
                        full_p = os.path.join(s_dir, f).replace('\\', '/')
                        if full_p in seen_paths or (template_path and os.path.abspath(full_p) == os.path.abspath(template_path)):
                            continue
                        seen_paths.add(full_p)

                        # Быстрая проверка по имени файла
                        is_matched = any(cls.is_house_match(tgt, f) for tgt in match_targets)

                        # Если по имени не подошло, проверяем только кандидатов с именами вроде 123.xlsx, 1234.xlsx или аркус
                        if not is_matched:
                            f_clean = f.lower()
                            is_arc_candidate = (
                                re.match(r'^\d+\.xlsx?$', f_clean)
                                or any(k in f_clean for k in ['аркус', 'arcus', 'выгруз', 'показан', '1с', 'свод', 'отчет'])
                            )
                            if is_arc_candidate:
                                arc_addr = cls.extract_house_name_from_arcus_content(full_p)
                                if arc_addr:
                                    is_matched = any(cls.is_house_match(tgt, arc_addr) for tgt in match_targets)

                        if is_matched:
                            f_low = f.lower()
                            dir_low = s_dir.lower()

                            # Файл Аркуса должен относиться к следующему месяцу
                            is_target_month = (
                                next_m_name in dir_low
                                or next_m_str in dir_low
                                or (next_month_dir and os.path.commonpath([s_dir, next_month_dir]) == next_month_dir)
                            )
                            # Если в пути явно указан текущий или старый месяц и нет целевого - отсекаем
                            if not is_target_month or (curr_m_name in dir_low and next_m_name not in dir_low):
                                continue

                            score = 100
                            # Если в папке целевого месяца
                            if next_m_name in dir_low or next_m_str in dir_low:
                                score += 60
                            # Если в названии файла/папки есть аркус/выгрузка
                            if any(k in f_low for k in ['аркус', 'arcus', 'выгруз']):
                                score += 50
                            if any(k in dir_low for k in ['аркус', 'arcus', 'выгруз']):
                                score += 40
                            # Если в целевой папке следующего месяца
                            if next_month_dir and os.path.commonpath([s_dir, next_month_dir]) == next_month_dir:
                                score += 40

                            # Свежесть файла (бонус по времени изменения)
                            try:
                                mtime = os.path.getmtime(full_p)
                                score += (mtime / 1e12)
                            except Exception:
                                pass

                            candidates.append((score, full_p))
            except Exception:
                pass

        if candidates:
            candidates.sort(key=lambda x: x[0], reverse=True)
            return candidates[0][1], next_month_dir

        return None, next_month_dir

    @classmethod
    def validate_arcus_month_folder(cls, template_path: str, arcus_path: str) -> Dict:
        """
        Проверяет, что файл Аркус взят строго из папки следующего (расчетного) месяца.
        Если папка не совпадает, возвращает флаг ошибки и поясняющее предупреждение.
        """
        res = {
            "is_valid": True,
            "target_month": None,
            "target_year": None,
            "target_month_name": "",
            "arcus_month": None,
            "arcus_year": None,
            "arcus_month_name": "",
            "warning": ""
        }
        if not template_path or not arcus_path:
            return res

        m_tpl, y_tpl = cls.extract_month_and_year_from_path(template_path)
        if not m_tpl:
            now = datetime.now()
            m_tpl, y_tpl = now.month, now.year

        next_m = 1 if m_tpl == 12 else m_tpl + 1
        next_y = y_tpl + 1 if m_tpl == 12 else y_tpl
        next_m_name = MONTH_NAMES_CAP[next_m - 1]

        res["target_month"] = next_m
        res["target_year"] = next_y
        res["target_month_name"] = f"{next_m_name} {next_y}"

        m_arc, y_arc = cls.extract_month_and_year_from_path(arcus_path)
        if m_arc:
            arc_m_name = MONTH_NAMES_CAP[m_arc - 1]
            res["arcus_month"] = m_arc
            res["arcus_year"] = y_arc
            res["arcus_month_name"] = f"{arc_m_name} {y_arc}" if y_arc else arc_m_name

            if m_arc != next_m or (y_arc and y_arc != next_y):
                res["is_valid"] = False
                res["warning"] = (
                    f"Файл Аркуса взят из папки «{res['arcus_month_name']}», "
                    f"но расчет производится на «{res['target_month_name']}»! "
                    f"Файл Аркус должен браться строго из папки следующего месяца."
                )
        else:
            # Проверяем вхождение названий месяцев в путь к файлу Аркус
            norm_arc = arcus_path.lower().replace('\\', '/')
            if MONTHS_RU[m_tpl - 1] in norm_arc and MONTHS_RU[next_m - 1] not in norm_arc:
                res["is_valid"] = False
                curr_m_name = MONTH_NAMES_CAP[m_tpl - 1]
                res["arcus_month_name"] = f"{curr_m_name} {y_tpl}"
                res["warning"] = (
                    f"Файл Аркуса находится в папке прошлого месяца «{curr_m_name} {y_tpl}»! "
                    f"Для расчета на «{next_m_name} {next_y}» требуется файл Аркус из папки следующего месяца."
                )

        return res

    @classmethod
    def detect_folder_context(cls, template_path: str) -> Dict:
        """
        Анализирует контекст шаблона:
        - Исходную папку и корень проекта (напр. 'Южный город')
        - Месяц шаблона и целевой расчетный месяц
        - Папку следующего месяца
        - Рекомендуемый путь сохранения (включая папку ГОТОВО)
        - Автоматически найденный файл Аркус в целевом месяце (поиск на 2-3 папки вверх)
        - Статус существующего отчета
        """
        result = {
            "template_path": template_path,
            "root_dir": None,
            "current_month_dir": None,
            "current_month": None,
            "current_year": None,
            "next_month": None,
            "next_year": None,
            "next_month_name": None,
            "next_month_dir": None,
            "found_arcus_path": None,
            "suggested_save_path": None,
            "existing_report_path": None,
            "existing_report_info": None,
            "house_name": None,
            "target_filename": None,
            "available_houses": []
        }

        if not template_path or not os.path.exists(template_path):
            return result

        from core.excel_parser import ExcelManager

        if os.path.isdir(template_path):
            current_month_dir = os.path.abspath(template_path)
            if os.path.basename(current_month_dir).upper() in ("ГОТОВО", "DONE", "READY"):
                current_month_dir = os.path.dirname(current_month_dir)
            tpl_filename = ""
            clean_house = ""
        else:
            tpl_dir = os.path.dirname(os.path.abspath(template_path))
            if os.path.basename(tpl_dir).upper() in ("ГОТОВО", "DONE", "READY"):
                current_month_dir = os.path.dirname(tpl_dir)
            else:
                current_month_dir = tpl_dir

            tpl_filename = os.path.basename(template_path)
            # Извлекаем точное название дома из содержимого шаблона (строка 2 таблицы)
            extracted_house = ExcelManager.extract_house_name(template_path)
            clean_house = extracted_house if extracted_house else tpl_filename.replace('.xlsx', '').replace('.xls', '').replace('+', '').strip()

        root_dir = os.path.dirname(current_month_dir)
        result["current_month_dir"] = current_month_dir
        result["root_dir"] = root_dir

        m, y = cls.extract_month_and_year_from_path(template_path)
        if not m:
            now = datetime.now()
            m, y = now.month, now.year

        result["current_month"] = m
        result["current_year"] = y

        # Расчет следующего месяца
        next_m = 1 if m == 12 else m + 1
        next_y = y + 1 if m == 12 else y
        result["next_month"] = next_m
        result["next_year"] = next_y
        result["next_month_name"] = MONTH_NAMES_CAP[next_m - 1]

        # Список всех доступных домов в текущей папке
        houses = []
        if os.path.exists(current_month_dir):
            for f in sorted(os.listdir(current_month_dir)):
                if f.endswith('.xlsx') and not f.startswith('~$') and not any(ex in f.lower() for ex in ['сопроводит', 'акт', 'аркус']):
                    full_p = os.path.join(current_month_dir, f).replace('\\', '/')
                    h_name = f.replace('.xlsx', '').replace('.xls', '').replace('+', '').strip()
                    houses.append({
                        "name": h_name,
                        "filename": f,
                        "path": full_p
                    })
        result["available_houses"] = houses

        if not clean_house and houses:
            clean_house = houses[0]["name"]
            tpl_filename = houses[0]["filename"]
            result["template_path"] = houses[0]["path"]

        result["house_name"] = clean_house

        # Умный поиск Аркуса и папки следующего месяца на 2-3 папки вверх
        found_arc, next_month_dir = cls.find_smart_arcus_path(
            tpl_filename=tpl_filename,
            current_month_dir=current_month_dir,
            next_m=next_m,
            next_y=next_y,
            curr_m=m,
            template_path=template_path,
            house_name=clean_house
        )

        result["next_month_dir"] = next_month_dir
        result["found_arcus_path"] = found_arc

        # 3. Формируем рекомендуемый путь сохранения и проверяем существующий отчет
        target_dir = next_month_dir if next_month_dir else current_month_dir
        ready_subdir = os.path.join(target_dir, "ГОТОВО")
        save_folder = ready_subdir if os.path.isdir(ready_subdir) else target_dir

        out_fn = ExcelManager.parse_house_and_next_month(template_path) if template_path and os.path.isfile(template_path) else (f"{clean_house}.xlsx" if clean_house else "Отчет.xlsx")
        target_filename = out_fn
        result["target_filename"] = target_filename
        suggested_save = os.path.join(save_folder, target_filename).replace('\\', '/')
        result["suggested_save_path"] = suggested_save

        possible_exist_paths = [
            suggested_save,
            os.path.join(target_dir, target_filename).replace('\\', '/'),
            os.path.join(target_dir, "ГОТОВО", target_filename).replace('\\', '/'),
            os.path.join(target_dir, "Готово", target_filename).replace('\\', '/')
        ]

        for p_chk in possible_exist_paths:
            if os.path.isfile(p_chk) and os.path.abspath(p_chk) != os.path.abspath(template_path):
                result["existing_report_path"] = p_chk
                try:
                    mtime = os.path.getmtime(p_chk)
                    dt_str = datetime.fromtimestamp(mtime).strftime("%d.%m в %H:%M")
                    sz_kb = round(os.path.getsize(p_chk) / 1024, 1)
                    result["existing_report_info"] = f"{os.path.basename(p_chk)} ({sz_kb} KB, изменен {dt_str})"
                except Exception:
                    result["existing_report_info"] = os.path.basename(p_chk)
                break

        return result

    @classmethod
    def check_file_locked_by_excel(cls, file_path: str) -> bool:
        """
        Проверяет, открыт ли файл в другой программе (Excel).
        Возвращает True, если файл заблокирован.
        """
        if not file_path or not os.path.exists(file_path):
            return False

        try:
            with open(file_path, 'r+b'):
                return False
        except (PermissionError, IOError):
            return True
        except Exception:
            return False

    @classmethod
    def generate_safe_copy_path(cls, save_path: str) -> str:
        """Генерирует альтернативный безопасный путь файла (например 'Душистая 45_v2.xlsx')."""
        if not save_path:
            return "Отчет_копия.xlsx"

        dirname = os.path.dirname(save_path)
        base = os.path.basename(save_path)
        name, ext = os.path.splitext(base)

        idx = 2
        while True:
            candidate = os.path.join(dirname, f"{name}_v{idx}{ext}").replace('\\', '/')
            if not os.path.exists(candidate) and not cls.check_file_locked_by_excel(candidate):
                return candidate
            idx += 1
            if idx > 99:
                return os.path.join(dirname, f"{name}_{int(time.time())}{ext}").replace('\\', '/')

    @classmethod
    def cross_validate_apartments(cls, template_path: str, arcus_path: str, excel_manager=None) -> Dict:
        """
        Сравнивает пересечение квартир между Шаблоном и Аркусом.
        Возвращает метрики совпадения и предупреждения при несовместимости.
        """
        res = {
            "is_valid": True,
            "match_pct": 100.0,
            "matched_count": 0,
            "tpl_count": 0,
            "arc_count": 0,
            "warning": ""
        }

        if not template_path or not arcus_path or not os.path.exists(template_path) or not os.path.exists(arcus_path):
            return res

        try:
            import openpyxl
            wb_tpl = openpyxl.load_workbook(template_path, data_only=True)
            ws_tpl = wb_tpl.active
            wb_arc = openpyxl.load_workbook(arcus_path, data_only=True)
            ws_arc = wb_arc.active

            def get_apts(ws):
                apts = set()
                for r in range(1, min(ws.max_row + 1, 300)):
                    val = str(ws.cell(row=r, column=1).value or '').strip().lower()
                    if val and not any(ex in val for ex in ['итого', 'всего', 'закрытые', 'замененные', 'директор', 'подпись', 'реестр']):
                        m_num = re.search(r'\d+', val)
                        if m_num:
                            apts.add(m_num.group(0))
                return apts

            tpl_apts = get_apts(ws_tpl)
            arc_apts = get_apts(ws_arc)

            res["tpl_count"] = len(tpl_apts)
            res["arc_count"] = len(arc_apts)

            if tpl_apts and arc_apts:
                intersection = tpl_apts.intersection(arc_apts)
                res["matched_count"] = len(intersection)
                pct = (len(intersection) / len(tpl_apts)) * 100.0
                res["match_pct"] = round(pct, 1)

                if pct < 20.0:
                    res["is_valid"] = False
                    res["warning"] = f"Критическое несовпадение квартир! Совпало всего {len(intersection)} из {len(tpl_apts)} ({pct:.1f}%). Возможно, выбран файл другого дома!"
                elif pct < 70.0:
                    res["warning"] = f"Внимание: совпало {len(intersection)} из {len(tpl_apts)} квартир ({pct:.1f}%). Проверьте корректность выгрузки Аркус."
        except Exception as e:
            res["warning"] = f"Не удалось выполнить предварительную сверку: {e}"

        return res
